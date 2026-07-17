"""4G 日监控 - 零低流量风险小区分析。

输入：data/ 下「重要场景-天*.xlsx」（可多个，按日期合并）
可选：data/ 或项目根下「问题小区问题归类.xlsx」
输出：项目根「4G日监控_零低流量风险小区_<监控日期>.xlsx」
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from app.pipelines.core import (
    BASE_DIR,
    DATA_DIR,
    GuiLogger,
    GuiProgress,
    LogCallback,
    ProgressCallback,
    SourceFileError,
)

LOW_THRESHOLD_GB = 0.1  # 低流量阈值：100MB = 0.1GB
OUTPUT_PREFIX = "4G日监控_零低流量风险小区"
DAY_FLOW_PATTERN = "重要场景-天*.xlsx"
PROBLEM_FILE_NAMES = ("问题小区问题归类.xlsx",)

RISK_COLORS = {
    "严重": ("C00000", "FFFFFF"),
    "高危": ("FF0000", "FFFFFF"),
    "中危": ("E36C0A", "FFFFFF"),
    "预警": ("FFC000", "000000"),
    "关注": ("FFFF00", "000000"),
    "正常": ("FFFFFF", "000000"),
}

RISK_ORDER = {"严重": 0, "高危": 1, "中危": 2, "预警": 3, "关注": 4, "正常": 5}


def find_day_flow_files(data_dir: Path | None = None) -> list[Path]:
    root = data_dir or DATA_DIR
    files = sorted(
        p for p in root.glob(DAY_FLOW_PATTERN) if p.is_file() and not p.name.startswith(".~")
    )
    if not files:
        raise SourceFileError(f"未找到天流量文件: {root / DAY_FLOW_PATTERN}")
    return files


def find_problem_file(data_dir: Path | None = None) -> Path | None:
    roots = [data_dir or DATA_DIR, BASE_DIR]
    for root in roots:
        for name in PROBLEM_FILE_NAMES:
            path = root / name
            if path.is_file():
                return path
    return None


def load_and_merge(file_paths: list[Path], logger: GuiLogger) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for fp in file_paths:
        xl = pd.ExcelFile(fp)
        sheet = xl.sheet_names[0]
        df = pd.read_excel(fp, sheet_name=sheet)
        df["__源文件__"] = fp.name
        frames.append(df)
        cgi_n = df["CGI"].nunique() if "CGI" in df.columns else 0
        logger.log(f"  加载: {fp.name} => {len(df)} 行, {cgi_n} 个小区")
    return pd.concat(frames, ignore_index=True)


def load_problems(problem_file: Path | None, logger: GuiLogger) -> pd.DataFrame | None:
    if problem_file is None or not problem_file.is_file():
        logger.log("[警告] 未找到问题小区归类文件，跳过关联")
        return None
    xl = pd.ExcelFile(problem_file)
    df = pd.read_excel(problem_file, sheet_name=xl.sheet_names[0])
    if "CGI" not in df.columns or "问题" not in df.columns:
        logger.log(f"[警告] 问题文件缺少 CGI/问题 列: {problem_file.name}")
        return None
    df = df.copy()
    df["问题"] = df["问题"].apply(
        lambda x: "" if pd.isna(x) or str(x).strip() == "0" else str(x).strip()
    )
    logger.log(
        f"  加载问题小区: {problem_file.name} => {len(df)} 行, "
        f"有效问题数: {(df['问题'] != '').sum()}"
    )
    return df[["CGI", "问题"]]


def calc_consecutive(row: pd.Series, date_cols: list[str], mode: str = "zero") -> int:
    count = 0
    for d in reversed(date_cols):
        v = row.get(d)
        if pd.isna(v):
            break
        if mode == "zero":
            if v == 0:
                count += 1
            else:
                break
        elif mode == "low":
            if v > 0 and v < LOW_THRESHOLD_GB:
                count += 1
            else:
                break
    return count


def risk_level(zd: int, ld: int) -> str:
    if zd >= 7:
        return "严重"
    if zd >= 5 or ld >= 7:
        return "高危"
    if zd >= 3 or ld >= 5:
        return "中危"
    if zd >= 1 or ld >= 3:
        return "预警"
    if ld >= 1:
        return "关注"
    return "正常"


def _style_header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="366092")
    cell.alignment = Alignment(horizontal="center")


def write_report(
    pivot: pd.DataFrame,
    dates: list,
    monitor_date,
    summary_stats: dict[str, Any],
    output_file: Path,
) -> None:
    wb = Workbook()
    date_cols = [str(d) for d in dates]
    risk_counts = summary_stats["risk_counts"]

    # Sheet1: 监控汇总
    ws1 = wb.active
    ws1.title = "监控汇总"
    summary_rows = [
        ["监控指标", "数值"],
        ["监控日期", str(monitor_date)],
        ["数据范围", f"{dates[0]} 至 {dates[-1]}（共{len(dates)}天）"],
        ["小区总数", summary_stats["total"]],
        ["当日零流量小区数", summary_stats["zero_today"]],
        ["当日低流量小区数", summary_stats["low_today"]],
        ["当日零低流量小区数", summary_stats["zero_low_today"]],
        ["当日零低流量占比", f"{summary_stats['ratio']:.4%}"],
        ["已关联已知问题小区数", summary_stats["with_problem"]],
        ["", ""],
        ["风险等级", "小区数"],
    ]
    for lv in ["严重", "高危", "中危", "预警", "关注", "正常"]:
        summary_rows.append([lv, risk_counts.get(lv, 0)])

    for r_idx, row in enumerate(summary_rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws1.cell(r_idx, c_idx, val)
            if r_idx == 1 or (r_idx == 11 and c_idx == 1):
                _style_header(cell)
            elif c_idx == 1:
                cell.font = Font(bold=True)
            elif r_idx == 8 and c_idx == 2:
                cell.font = Font(bold=True, color="C00000")
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 22

    # Sheet2: 风险小区明细
    ws2 = wb.create_sheet("风险小区明细")
    risk_df = pivot[pivot["当日状态"].isin(["零流量", "低流量"])].copy()
    risk_df["_risk_rank"] = risk_df["风险等级"].map(RISK_ORDER).fillna(99)
    risk_df = risk_df.sort_values(["_risk_rank", "连续零低流量天数"], ascending=[True, False])

    out_cols = [
        "CGI",
        "小区名称",
        "所属地市",
        "当日流量_GB",
        "当日状态",
        "连续零流量天数",
        "连续低流量天数",
        "连续零低流量天数",
        "风险等级",
        "问题",
    ]
    risk_out = risk_df[out_cols].copy()
    risk_out.columns = [
        "CGI",
        "小区名称",
        "所属地市",
        "当日流量(GB)",
        "当日状态",
        "连续零流量天数",
        "连续低流量天数",
        "连续零低流量天数",
        "风险等级",
        "已知问题",
    ]

    for r_idx, row in enumerate(dataframe_to_rows(risk_out, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws2.cell(r_idx, c_idx, val)
            if r_idx == 1:
                _style_header(cell)
            else:
                if c_idx == 9:
                    bg, fg = RISK_COLORS.get(val, ("FFFFFF", "000000"))
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(bold=True, color=fg)
                elif c_idx == 10 and val:
                    cell.font = Font(color="C00000")
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
                elif r_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F7F9FC")

    for col in ["A", "B", "C"]:
        ws2.column_dimensions[col].width = 26
    for col in ["D", "E", "F", "G", "H", "I"]:
        ws2.column_dimensions[col].width = 16
    ws2.column_dimensions["J"].width = 40

    # Sheet3: 全量监控明细
    ws3 = wb.create_sheet("全量监控明细")
    all_out = pivot[
        ["CGI", "小区名称", "所属地市"]
        + date_cols
        + [
            "当日流量_GB",
            "当日状态",
            "连续零流量天数",
            "连续低流量天数",
            "连续零低流量天数",
            "风险等级",
            "问题",
        ]
    ].copy()
    all_out.columns = (
        ["CGI", "小区名称", "所属地市"]
        + [f"{d}流量(GB)" for d in dates]
        + [
            "当日流量(GB)",
            "当日状态",
            "连续零流量天数",
            "连续低流量天数",
            "连续零低流量天数",
            "风险等级",
            "已知问题",
        ]
    )

    risk_col_idx = len(all_out.columns) - 1
    problem_col_idx = len(all_out.columns)

    for r_idx, row in enumerate(dataframe_to_rows(all_out, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws3.cell(r_idx, c_idx, val)
            if r_idx == 1:
                _style_header(cell)
            else:
                if c_idx == risk_col_idx:
                    bg, fg = RISK_COLORS.get(val, ("FFFFFF", "000000"))
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(bold=True, color=fg)
                elif c_idx == problem_col_idx and val:
                    cell.font = Font(color="C00000")
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
                elif r_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F7F9FC")

    for col in ["A", "B", "C"]:
        ws3.column_dimensions[col].width = 26
    for c in range(4, 4 + len(dates)):
        ws3.column_dimensions[get_column_letter(c)].width = 16
    for c in range(4 + len(dates), problem_col_idx + 1):
        ws3.column_dimensions[get_column_letter(c)].width = 16
    ws3.column_dimensions[get_column_letter(problem_col_idx)].width = 40

    wb.save(output_file)


def run_zero_low_flow_pipeline(
    progress_callback: ProgressCallback | None = None,
    log_callback: LogCallback | None = None,
    data_dir: Path | str | None = None,
    output_dir: Path | str | None = None,
) -> Path:
    """运行零低流量风险小区分析，返回输出 Excel 路径。"""
    logger = GuiLogger(log_callback)
    progress = GuiProgress(progress_callback, logger)
    data_root = Path(data_dir) if data_dir else DATA_DIR
    out_root = Path(output_dir) if output_dir else BASE_DIR

    logger.log("=" * 50)
    logger.log("4G日监控 - 零低流量风险小区分析")
    logger.log("=" * 50)

    progress.update(5, f"读取数据目录: {data_root}")
    file_paths = find_day_flow_files(data_root)
    raw_df = load_and_merge(file_paths, logger)
    required = ["CGI", "小区名称", "所属地市", "记录开始时间", "日4G流量（GB）"]
    missing = [c for c in required if c not in raw_df.columns]
    if missing:
        raise SourceFileError(f"天流量文件缺少必要列: {', '.join(missing)}")
    logger.log(f"合并完成: {len(raw_df)} 行, {raw_df['CGI'].nunique()} 个唯一小区")

    progress.update(20, "读取问题小区归类...")
    problem_df = load_problems(find_problem_file(data_root), logger)

    progress.update(35, "数据预处理...")
    raw_df = raw_df.copy()
    raw_df["日期"] = pd.to_datetime(raw_df["记录开始时间"]).dt.date
    raw_df["流量_GB"] = pd.to_numeric(raw_df["日4G流量（GB）"], errors="coerce")
    dates = sorted(raw_df["日期"].unique())
    if not dates:
        raise SourceFileError("天流量数据中没有有效日期")
    monitor_date = dates[-1]
    output_file = out_root / f"{OUTPUT_PREFIX}_{monitor_date}.xlsx"
    logger.log(f"数据日期范围: {dates[0]} 至 {dates[-1]}, 共 {len(dates)} 天")
    logger.log(f"监控日期(最新): {monitor_date}")

    progress.update(55, "构建透视表并计算风险...")
    pivot = (
        raw_df.pivot_table(
            index=["CGI", "小区名称", "所属地市"],
            columns="日期",
            values="流量_GB",
            aggfunc="first",
        )
        .reset_index()
    )
    pivot.columns.name = None
    date_cols = [str(d) for d in dates]
    for d in dates:
        pivot = pivot.rename(columns={d: str(d)})

    pivot["当日流量_GB"] = pivot[str(monitor_date)]
    pivot["当日状态"] = pivot["当日流量_GB"].apply(
        lambda x: "零流量"
        if pd.isna(x) or x == 0
        else ("低流量" if x < LOW_THRESHOLD_GB else "正常")
    )
    pivot["连续零流量天数"] = pivot.apply(
        lambda r: calc_consecutive(r, date_cols, "zero"), axis=1
    )
    pivot["连续低流量天数"] = pivot.apply(
        lambda r: calc_consecutive(r, date_cols, "low"), axis=1
    )
    pivot["连续零低流量天数"] = pivot[["连续零流量天数", "连续低流量天数"]].max(axis=1)
    pivot["风险等级"] = pivot.apply(
        lambda r: risk_level(r["连续零流量天数"], r["连续低流量天数"]), axis=1
    )

    if problem_df is not None:
        pivot = pivot.merge(problem_df, on="CGI", how="left")
        pivot["问题"] = pivot["问题"].fillna("")
    else:
        pivot["问题"] = ""

    total = len(pivot)
    zero_today = int((pivot["当日状态"] == "零流量").sum())
    low_today = int((pivot["当日状态"] == "低流量").sum())
    zero_low_today = zero_today + low_today
    ratio = zero_low_today / total if total else 0.0
    risk_counts = pivot["风险等级"].value_counts().to_dict()
    with_problem = int((pivot["问题"] != "").sum())

    logger.log(f"小区总数: {total}")
    logger.log(f"当日零流量: {zero_today}, 当日低流量: {low_today}, 合计: {zero_low_today}")
    logger.log(f"当日零低流量占比: {ratio:.4%}")
    logger.log(f"风险分布: {dict(sorted(risk_counts.items(), key=lambda kv: -kv[1]))}")
    logger.log(f"已关联已知问题: {with_problem} 个小区")

    progress.update(85, "生成监控报表...")
    summary_stats = {
        "total": total,
        "zero_today": zero_today,
        "low_today": low_today,
        "zero_low_today": zero_low_today,
        "ratio": ratio,
        "with_problem": with_problem,
        "risk_counts": risk_counts,
    }
    write_report(pivot, dates, monitor_date, summary_stats, output_file)
    logger.log(f"报表已保存: {output_file.name}")
    progress.update(100, f"已生成: {output_file.name}")
    return output_file


def latest_zero_low_flow_output() -> Path | None:
    files = sorted(
        BASE_DIR.glob(f"{OUTPUT_PREFIX}_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return files[0] if files else None


__all__ = [
    "DAY_FLOW_PATTERN",
    "OUTPUT_PREFIX",
    "find_day_flow_files",
    "latest_zero_low_flow_output",
    "run_zero_low_flow_pipeline",
]
