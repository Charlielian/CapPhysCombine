from __future__ import annotations

import logging
import os
import queue
import re
import subprocess
import sys
import threading
import time
import traceback
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable, Iterable

import duckdb
import geopandas as gpd
import numpy as np
import pandas as pd
from shapely.geometry import Point

# ==============================================================================
# 物理表模块常量定义
# ==============================================================================

# 5G频点映射: {BAND: {BAND_A: 频点}}
NR_FREQ_MAPPING = {
    "2.6GHz": {"2.6G一载波": 504990, "2.6G二载波": 524910},
    "4.9GHz": {"4.9G一载波": 721824},
    "700M": {"700M": 152650},
}

# 4G BAND_A -> BAND
LTE_BAND_MAPPING = {
    "F1": "F",
    "F2": "F",
    "FDD1800": "FDD1800",
    "A频段": "A",
    "FDD900": "FDD900",
    "E1": "E",
    "E2": "E",
    "E3": "E",
    "D1": "D",
    "D3": "D",
    "D7": "D",
    "D8": "D",
    "NB": "NB",
}

BAND_3DMIMO = {"D1", "D3", "D7", "D8"}

LTE_FREQ_MAPPING = {
    "F1": 38400,
    "F2": 38544,
    "FDD1800": [1300, 1301],
    "A频段": 36275,
    "FDD900": 3590,
    "E1": 38950,
    "E2": 39148,
    "E3": 39292,
    "D1": 40936,
    "D3": 40936,
    "D7": 41134,
    "D8": 41332,
    "NB": None,
}

LTE_BANDS = {"F", "FDD1800", "FDD900", "D", "E", "A", "NB", "5G-3Dmimo"}
NR_BANDS = {"700M", "2.6GHz", "4.9GHz"}

DISTANCE_INDOOR_M = 100
DISTANCE_MACRO_M = 50

COVERAGE_LAYER_MAP = {
    "True": 1,
    "1": 1,
    "是": 1,
    "Yes": 1,
    "False": 0,
    "0": 0,
    "否": 0,
    "No": 0,
}

PHYSICAL_TABLE_AVAILABLE = True


def _get_base_dir() -> Path:
    """返回项目根目录（兼容 PyInstaller 打包和源码运行）。"""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    # app/pipelines/core.py -> project root
    return Path(__file__).resolve().parent.parent.parent


BASE_DIR = _get_base_dir()
DATA_DIR = BASE_DIR / "data"
LOG_DIR = BASE_DIR / "logs"
LOG_RETENTION_DAYS = 7
UNIFIED_DB_PATH = BASE_DIR / "capphys_unified.db"


# ==============================================================================
# 物理表模块：频段映射 (band.py)
# ==============================================================================

def map_lte_band(band_a: str, cell_name: str) -> str:
    """4G 详细频段 -> BAND；D 频段且小区名含 RD-Z 为 5G-3Dmimo。"""
    if band_a in BAND_3DMIMO and "RD-Z" in str(cell_name):
        return "5G-3Dmimo"
    return LTE_BAND_MAPPING.get(band_a, band_a)


def build_band_aggregations(df: pd.DataFrame, group_col: str, lte_bands=None):
    """按分组列汇总 BAND / LTE BAND 字符串。"""

    lte_bands = lte_bands or LTE_BANDS

    def calc_bands(x):
        bands = {str(b) for b in x if b and str(b) != "nan"}
        return "/".join(sorted(bands))

    def calc_lte_bands(x):
        lte_set = {str(b) for b in x if b in lte_bands}
        return "/".join(sorted(lte_set))

    agg_all = df.groupby(group_col)["BAND"].apply(calc_bands).to_dict()
    agg_lte = df.groupby(group_col)["BAND"].apply(calc_lte_bands).to_dict()
    return agg_all, agg_lte


def co_site_coverage_type(station_bands_str: str) -> str:
    """根据物理站制式字符串判断共站制式情况。"""
    if not station_bands_str:
        return ""
    bands = set(station_bands_str.split("/"))
    has_nr = bool(bands & NR_BANDS)
    has_lte = bool(bands & LTE_BANDS)
    if has_nr and has_lte:
        return "有5有4"
    if has_nr and not has_lte:
        return "有5无4"
    if not has_nr and has_lte:
        return "无5有4"
    return ""


def apply_lte_network_structure(agg_df: pd.DataFrame) -> pd.DataFrame:
    """向量化计算 4G 网络结构列。"""
    df = agg_df.copy()
    df["LTE频段数"] = df["物理扇区LTE制式"].str.count("/") + (
        df["物理扇区LTE制式"] != ""
    ).astype(int)
    df["覆盖层值"] = (
        df["覆盖层"].astype(str).map(COVERAGE_LAYER_MAP).fillna(-1).astype(int)
    )
    conditions = [
        (df["覆盖层值"] == 1) & (df["LTE频段数"] == 1),
        (df["覆盖层值"] == 0) & (df["LTE频段数"] == 1),
        (df["覆盖层值"] == 1) & (df["LTE频段数"] > 1),
        (df["覆盖层值"] == 0) & (df["LTE频段数"] > 1),
    ]
    choices = ["单层网_覆盖层", "单层网", "多层网_覆盖层", "多层网"]
    df["网络结构4G"] = np.select(conditions, choices, default="")
    df.drop(columns=["LTE频段数", "覆盖层值"], inplace=True)
    return df


# ==============================================================================
# 物理表模块：扇区解析 (section.py)
# ==============================================================================

_SECTIONID_PATTERN = re.compile(r"扇区(\d+)")


def extract_sectionid(name):
    """从共站同覆盖名中提取扇区编号，如 xxx-扇区2 -> 2。"""
    if not name:
        return None
    match = _SECTIONID_PATTERN.search(str(name))
    if match:
        return int(match.group(1))
    return None


# ==============================================================================
# 物理表模块：空间查询 (spatial.py)
# ==============================================================================

def load_geojson_with_index(file_path):
    """加载 GeoJSON 并返回 (gdf, spatial_index)，文件不存在或为空则索引为 None。"""
    if not os.path.exists(file_path):
        return None, None
    gdf = gpd.read_file(file_path)
    if len(gdf) == 0:
        return gdf, None
    return gdf, gdf.sindex


def _is_invalid_coord(lon, lat):
    if pd.isna(lon) or pd.isna(lat):
        return True
    if lon == 0 or lat == 0 or str(lon) == "***":
        return True
    return False


def get_grid_by_coords_batch(gdf, sindex, lons, lats):
    """批量根据经纬度做点在多边形内查询，返回每点匹配的 GeoSeries 行或 None。"""
    if gdf is None or sindex is None:
        return [None] * len(lons)

    results = [None] * len(lons)
    for i in range(len(lons)):
        lon, lat = lons[i], lats[i]
        if _is_invalid_coord(lon, lat):
            continue
        try:
            point = Point(float(lon), float(lat))
            for idx in sindex.intersection(point.bounds):
                if gdf.iloc[idx].geometry.contains(point):
                    results[i] = gdf.iloc[idx]
                    break
        except (ValueError, TypeError):
            pass
    return results


# ==============================================================================
# 物理表模块：距离聚类 (clustering.py)
# ==============================================================================

def union_find_cluster(lons, lats, station_names, distance_threshold_m):
    """
    对 n 条记录做距离聚类；同一物理站名强制同簇。
    返回长度为 n 的 cluster_id 列表。
    """
    n = len(lons)
    if n == 0:
        return []

    parent = list(range(n))

    def find(x):
        if parent[x] != x:
            parent[x] = find(parent[x])
        return parent[x]

    def union(x, y):
        px, py = find(x), find(y)
        if px != py:
            parent[px] = py

    station_groups = defaultdict(list)
    for i in range(n):
        name = station_names[i] if i < len(station_names) else ""
        if name and str(name) != "nan":
            station_groups[str(name)].append(i)

    for group in station_groups.values():
        if len(group) > 1:
            for k in range(1, len(group)):
                union(group[0], group[k])

    threshold_deg = distance_threshold_m / 1000 / 111
    grid_size = threshold_deg
    grid_dict = defaultdict(list)

    for i in range(n):
        grid_x = int(lons[i] / grid_size)
        grid_y = int(lats[i] / grid_size)
        grid_dict[(grid_x, grid_y)].append(i)

    for i in range(n):
        grid_x = int(lons[i] / grid_size)
        grid_y = int(lats[i] / grid_size)
        for dx in (-1, 0, 1):
            for dy in (-1, 0, 1):
                key = (grid_x + dx, grid_y + dy)
                if key not in grid_dict:
                    continue
                for j in grid_dict[key]:
                    if i >= j or find(i) == find(j):
                        continue
                    dlat = np.radians(lats[j] - lats[i])
                    dlon = np.radians(lons[j] - lons[i])
                    a = (
                        np.sin(dlat / 2) ** 2
                        + np.cos(np.radians(lats[i]))
                        * np.cos(np.radians(lats[j]))
                        * np.sin(dlon / 2) ** 2
                    )
                    dist = 2 * 6371000 * np.arcsin(np.sqrt(a))
                    if dist <= distance_threshold_m:
                        union(i, j)

    cluster_map = {}
    for i in range(n):
        root = find(i)
        if root not in cluster_map:
            cluster_map[root] = len(cluster_map)

    return [cluster_map[find(i)] for i in range(n)]


def cluster_by_distance(df: pd.DataFrame, distance_threshold_m, station_type_label: str):
    """返回 cluster_labels, cluster_names（与 df 行对齐）。"""
    if len(df) == 0:
        return [], []

    lons = df["经度"].values.astype(float)
    lats = df["纬度"].values.astype(float)
    station_names = df["物理站"].values

    valid_mask = ~(np.isnan(lons) | np.isnan(lats) | (lons == 0) | (lats == 0))
    valid_indices = np.where(valid_mask)[0]

    cluster_labels = [-1] * len(df)
    cluster_names = [""] * len(df)

    if len(valid_indices) == 0:
        for i in range(len(df)):
            name = station_names[i]
            cluster_labels[i] = i
            cluster_names[i] = (
                str(name) if name and str(name) != "nan" else f"{station_type_label}_未知{i}"
            )
        return cluster_labels, cluster_names

    valid_lons = lons[valid_mask]
    valid_lats = lats[valid_mask]
    valid_stations = station_names[valid_mask]
    valid_labels = union_find_cluster(
        valid_lons, valid_lats, valid_stations, distance_threshold_m
    )

    cluster_station_counts = defaultdict(lambda: defaultdict(int))
    for i, label in enumerate(valid_labels):
        name = valid_stations[i]
        if name and str(name) != "nan":
            cluster_station_counts[label][str(name)] += 1
        else:
            cluster_station_counts[label]["__unknown__"] += 1

    cluster_names_dict = {}
    for cluster_id, station_counts in cluster_station_counts.items():
        max_count = 0
        max_name = f"{station_type_label}_未知{cluster_id}"
        for name, count in station_counts.items():
            if name != "__unknown__" and count > max_count:
                max_count = count
                max_name = name
        cluster_names_dict[cluster_id] = max_name

    for local_i, orig_idx in enumerate(valid_indices):
        cluster_labels[orig_idx] = valid_labels[local_i]
        cluster_names[orig_idx] = cluster_names_dict[valid_labels[local_i]]

    for i in range(len(df)):
        if cluster_labels[i] != -1:
            continue
        name = station_names[i]
        cluster_names[i] = (
            str(name) if name and str(name) != "nan" else f"{station_type_label}_未知{i}"
        )
        found = False
        for j in range(len(df)):
            if cluster_labels[j] != -1 and str(station_names[j]) == str(name):
                cluster_labels[i] = cluster_labels[j]
                cluster_names[i] = cluster_names[j]
                found = True
                break
        if not found:
            cluster_labels[i] = max(cluster_labels) + 1 if max(cluster_labels) >= 0 else 0
            cluster_names[i] = (
                str(name) if name and str(name) != "nan" else f"{station_type_label}_未知{i}"
            )

    return cluster_labels, cluster_names


def aggregate_by_distance(agg_df: pd.DataFrame) -> pd.DataFrame:
    """室分 100m、宏站 50m 距离聚合及制式汇总列。"""
    print("\n按距离聚合物理站...")
    df = agg_df.copy()
    df["物理站名_距离聚合"] = ""

    indoor_mask = (df["站点类型"] == "室分") | (df["覆盖类型"] == "室内")
    macro_mask = (df["站点类型"] == "宏站") & (df["覆盖类型"] == "室外")
    other_mask = ~indoor_mask & ~macro_mask

    print(
        f"室分数量: {indoor_mask.sum()}, 宏站数量: {macro_mask.sum()}, "
        f"其他数量: {other_mask.sum()}"
    )

    if indoor_mask.sum() > 0:
        indoor_df = df.loc[indoor_mask].reset_index(drop=True)
        _, indoor_names = cluster_by_distance(indoor_df, DISTANCE_INDOOR_M, "室分")
        for idx, orig_idx in enumerate(df.index[indoor_mask]):
            df.loc[orig_idx, "物理站名_距离聚合"] = indoor_names[idx]

    macro_mask_combined = macro_mask | other_mask
    if macro_mask_combined.sum() > 0:
        macro_df = df.loc[macro_mask_combined].reset_index(drop=True)
        _, macro_names = cluster_by_distance(macro_df, DISTANCE_MACRO_M, "宏站")
        for idx, orig_idx in enumerate(df.index[macro_mask_combined]):
            df.loc[orig_idx, "物理站名_距离聚合"] = macro_names[idx]

    station_all_agg, station_lte_agg = build_band_aggregations(
        df, "物理站名_距离聚合", LTE_BANDS
    )

    df["物理站LTE制式_距离聚合"] = (
        df["物理站名_距离聚合"].map(station_lte_agg).fillna("")
    )
    df["物理站制式_距离聚合"] = df["物理站名_距离聚合"].map(station_all_agg).fillna("")
    df["共站制式情况_距离聚合"] = df["物理站制式_距离聚合"].apply(co_site_coverage_type)

    print(f"距离聚合完成，共 {df['物理站名_距离聚合'].nunique()} 个聚合物理站")
    return df


# ==============================================================================
# 物理表模块：数据读取 (readers.py)
# ==============================================================================

def calc_nr_freq(band, band_a):
    if band in NR_FREQ_MAPPING and band_a in NR_FREQ_MAPPING[band]:
        return NR_FREQ_MAPPING[band][band_a]
    return None


def read_nr_cellant(file_path: str) -> pd.DataFrame:
    df = pd.read_excel(file_path)
    result = pd.DataFrame(
        {
            "网络制式": "5G",
            "CGI": df["CGI"].fillna(""),
            "小区名称": df["小区名称"].fillna(""),
            "物理站": df["所属局站"].fillna(""),
            "物理站ID": df["所属局站ID"].fillna(""),
            "经度": pd.to_numeric(df["经度"], errors="coerce").fillna(0),
            "纬度": pd.to_numeric(df["纬度"], errors="coerce").fillna(0),
            "方位角": pd.to_numeric(df["方位角"], errors="coerce").fillna(0),
            "天线名": df["天线名称"].fillna(""),
            "挂高": pd.to_numeric(df["挂高"], errors="coerce").fillna(0),
            "厂家": df["厂家"].fillna(""),
            "BAND": df["使用频段"].fillna(""),
            "BAND_A": df["详细使用频段"].fillna(""),
            "站点类型": df["站点类型"].fillna(""),
            "网元状态": df["网元状态"].fillna(""),
            "覆盖类型": df["覆盖类型"].fillna(""),
            "乡镇街道": df["乡镇街道"].fillna(""),
            "一级标签": df["一级标签"].fillna(""),
            "路测网格": df["路测网格"].fillna(""),
            "来源文件": os.path.basename(file_path),
        }
    )
    result["频点"] = result.apply(
        lambda row: calc_nr_freq(row["BAND"], row["BAND_A"]), axis=1
    )
    return result


def read_lte_cellant(file_path: str) -> pd.DataFrame:
    df = pd.read_excel(file_path)
    net_types = df["网络制式"].fillna("")
    cell_names = df["小区名称"].fillna("")
    band_a_list = df["详细使用频段"].fillna("")

    band_df = pd.DataFrame({"band_a": band_a_list, "cell_name": cell_names})
    bands = band_df.apply(
        lambda row: map_lte_band(row["band_a"], row["cell_name"]), axis=1
    )

    def calc_lte_freq(row):
        band_a = row["band_a"]
        net_type = row["net_type"]
        if band_a not in LTE_FREQ_MAPPING:
            return None
        freq_val = LTE_FREQ_MAPPING[band_a]
        if net_type == "TDD":
            if "中心载频的信道号" in df.columns:
                return pd.to_numeric(
                    df.loc[row.name, "中心载频的信道号"], errors="coerce"
                )
            return freq_val
        if net_type == "FDD":
            if "下行中心载频的信道号" in df.columns:
                return pd.to_numeric(
                    df.loc[row.name, "下行中心载频的信道号"], errors="coerce"
                )
            return freq_val
        if net_type == "NB-IoT":
            return None
        return freq_val

    freq_df = pd.DataFrame({"band_a": band_a_list, "net_type": net_types})
    freqs = freq_df.apply(calc_lte_freq, axis=1)

    result = pd.DataFrame(
        {
            "网络制式": "4G",
            "CGI": df["CGI"].fillna(""),
            "小区名称": cell_names,
            "物理站": df["所属站点名称"].fillna(""),
            "物理站ID": df["站点ID"].fillna(""),
            "经度": pd.to_numeric(df["经度"], errors="coerce").fillna(0),
            "纬度": pd.to_numeric(df["纬度"], errors="coerce").fillna(0),
            "方位角": pd.to_numeric(df["方位角"], errors="coerce").fillna(0),
            "天线名": df["天线名称"].fillna(""),
            "挂高": pd.to_numeric(df["挂高"], errors="coerce").fillna(0),
            "厂家": df["厂家"].fillna(""),
            "BAND": bands,
            "BAND_A": band_a_list,
            "频点": freqs,
            "站点类型": df["站点类型"].fillna(""),
            "网元状态": df["网元状态"].fillna(""),
            "覆盖类型": df["覆盖类型"].fillna(""),
            "乡镇街道": df["乡镇街道"].fillna(""),
            "一级标签": df["一级标签"].fillna(""),
            "路测网格": df["路测网格"].fillna(""),
            "来源文件": os.path.basename(file_path),
        }
    )
    nb_count = (result["BAND"] == "NB").sum()
    result = result[result["BAND"] != "NB"]
    if nb_count > 0:
        print(f"  已过滤 {nb_count} 条 NB 小区")
    return result


def read_common_coverage(file_path: str) -> pd.DataFrame:
    df = pd.read_excel(file_path)
    region = df["小区所属区域"].fillna("")
    region = region.where(region != "", df["小区所属区域类型"].fillna(""))
    # Ensure 覆盖层 is always string (handle boolean from Excel)
    coverage_layer = df["是否覆盖层"].apply(lambda x: "是" if x is True else ("否" if x is False else (str(x) if pd.notna(x) else "")))
    result = pd.DataFrame(
        {
            "物理站名": df["物理站名"].fillna(""),
            "CGI": df["CGI"].fillna(""),
            "小区名称": df["小区名称"].fillna(""),
            "共站同覆盖名": df["共站同覆盖名"].fillna(""),
            "使用频段": df["使用频段"].fillna(""),
            "覆盖层": coverage_layer,
            "小区所属区域": region,
            "路测网格": df["路测网格"].fillna(""),
            "经度": pd.to_numeric(df["经度"], errors="coerce").fillna(0),
            "纬度": pd.to_numeric(df["纬度"], errors="coerce").fillna(0),
        }
    )
    result["sectionid"] = result["共站同覆盖名"].apply(extract_sectionid)
    return result


# ==============================================================================
# 统一数据库模块：共站同覆盖表持久化 + 两个功能共享
# ==============================================================================

def get_unified_db_connection() -> duckdb.DuckDBPyConnection:
    """获取统一数据库连接"""
    return duckdb.connect(str(UNIFIED_DB_PATH))


def _cog_coverage_table_exists(conn: duckdb.DuckDBPyConnection) -> bool:
    row = conn.execute(
        """
        SELECT 1 FROM information_schema.tables
        WHERE table_name = '共站同覆盖小区表'
        LIMIT 1
        """
    ).fetchone()
    return row is not None


def init_unified_database(conn: duckdb.DuckDBPyConnection | None = None) -> duckdb.DuckDBPyConnection:
    """初始化统一数据库表结构（包含持久化的共站同覆盖表）"""
    close_conn = conn is None
    if conn is None:
        conn = get_unified_db_connection()

    # 共站同覆盖表 - 持久化存储，支持两个功能共享
    # 模板列：CGI, 共站同覆盖名, 物理站名, 小区名称, 使用频段, 是否覆盖层, 小区所属区域, 路测网格, 经度, 纬度, sectionid
    conn.execute("""
        CREATE TABLE IF NOT EXISTS 共站同覆盖小区表 (
            CGI TEXT PRIMARY KEY,
            共站同覆盖名 TEXT NOT NULL,
            物理站名 TEXT,
            小区名称 TEXT,
            使用频段 TEXT,
            是否覆盖层 TEXT,
            小区所属区域 TEXT,
            路测网格 TEXT,
            经度 DOUBLE,
            纬度 DOUBLE,
            sectionid INTEGER,
            创建时间 TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            更新时间 TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # 兼容旧物理表 schema：覆盖层 -> 是否覆盖层
    cols = {
        r[0]
        for r in conn.execute(
            "SELECT column_name FROM information_schema.columns WHERE table_name = '共站同覆盖小区表'"
        ).fetchall()
    }
    if "是否覆盖层" not in cols and "覆盖层" in cols:
        conn.execute('ALTER TABLE 共站同覆盖小区表 RENAME COLUMN "覆盖层" TO "是否覆盖层"')
    elif "是否覆盖层" not in cols:
        conn.execute('ALTER TABLE 共站同覆盖小区表 ADD COLUMN "是否覆盖层" TEXT')

    # 创建索引
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cc_cgi ON 共站同覆盖小区表(CGI)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cc_name ON 共站同覆盖小区表(共站同覆盖名)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cc_station ON 共站同覆盖小区表(物理站名)")

    if not _cog_coverage_table_exists(conn):
        raise RuntimeError(f"统一数据库初始化失败，表未创建: {UNIFIED_DB_PATH}")

    if close_conn:
        conn.close()

    return conn


class CogCoverageManager:
    """共站同覆盖表管理器（CRUD操作）"""
    
    def __init__(self, conn: duckdb.DuckDBPyConnection | None = None):
        self.conn = conn
        self._own_connection = conn is None
        if conn is None:
            self.conn = get_unified_db_connection()
        # 确保数据库表结构已初始化
        init_unified_database(self.conn)
    
    def close(self):
        """关闭连接"""
        if self._own_connection and self.conn:
            self.conn.close()
            self.conn = None
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
    
    def get_all(self, limit: int | None = None, offset: int = 0) -> pd.DataFrame:
        """获取所有记录"""
        query = "SELECT * FROM 共站同覆盖小区表 ORDER BY CGI"
        if limit:
            query += f" LIMIT {limit} OFFSET {offset}"
        return self.conn.execute(query).fetchdf()
    
    def get_by_cgi(self, cgi: str) -> pd.DataFrame:
        """根据CGI查询单条记录"""
        return self.conn.execute(
            "SELECT * FROM 共站同覆盖小区表 WHERE CGI = ?", [cgi]
        ).fetchdf()
    
    def search(self, keyword: str) -> pd.DataFrame:
        """模糊搜索（物理站名、小区名称、共站同覆盖名）"""
        return self.conn.execute("""
            SELECT * FROM 共站同覆盖小区表 
            WHERE 物理站名 LIKE ? OR 小区名称 LIKE ? OR 共站同覆盖名 LIKE ? OR CGI LIKE ?
            ORDER BY CGI
        """, [f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"]
        ).fetchdf()
    
    def add(self, record: dict) -> bool:
        """添加单条记录"""
        try:
            self.conn.execute("""
                INSERT INTO 共站同覆盖小区表 
                (CGI, 共站同覆盖名, 物理站名, 小区名称, 使用频段, 是否覆盖层, 小区所属区域, 路测网格, 经度, 纬度, sectionid)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, [
                record.get("CGI"), record.get("共站同覆盖名"), record.get("物理站名"), 
                record.get("小区名称"), record.get("使用频段"), record.get("是否覆盖层"),
                record.get("小区所属区域"), record.get("路测网格"), record.get("经度"),
                record.get("纬度"), record.get("sectionid")
            ])
            return True
        except Exception as e:
            print(f"添加记录失败: {e}")
            return False
    
    def update(self, cgi: str, record: dict) -> bool:
        """更新单条记录"""
        try:
            self.conn.execute("""
                UPDATE 共站同覆盖小区表 SET
                    共站同覆盖名 = ?,
                    物理站名 = ?,
                    小区名称 = ?,
                    使用频段 = ?,
                    是否覆盖层 = ?,
                    小区所属区域 = ?,
                    路测网格 = ?,
                    经度 = ?,
                    纬度 = ?,
                    sectionid = ?,
                    更新时间 = CURRENT_TIMESTAMP
                WHERE CGI = ?
            """, [
                record.get("共站同覆盖名"), record.get("物理站名"), record.get("小区名称"),
                record.get("使用频段"), record.get("是否覆盖层"), record.get("小区所属区域"),
                record.get("路测网格"), record.get("经度"), record.get("纬度"),
                record.get("sectionid"), cgi
            ])
            return True
        except Exception as e:
            print(f"更新记录失败: {e}")
            return False
    
    def delete(self, cgi: str) -> bool:
        """删除单条记录"""
        try:
            self.conn.execute("DELETE FROM 共站同覆盖小区表 WHERE CGI = ?", [cgi])
            return True
        except Exception as e:
            print(f"删除记录失败: {e}")
            return False
    
    def delete_many(self, cgis: list[str]) -> int:
        """批量删除"""
        if not cgis:
            return 0
        try:
            placeholders = ",".join(["?"] * len(cgis))
            result = self.conn.execute(
                f"DELETE FROM 共站同覆盖小区表 WHERE CGI IN ({placeholders})", cgis
            )
            return result.fetchone()[0] if result.fetchone() else 0
        except Exception as e:
            print(f"批量删除失败: {e}")
            return 0
    
    def import_from_excel(self, excel_path: Path | str, replace: bool = False) -> int:
        """从Excel导入共站同覆盖表 - 使用标准模板
        
        模板列：CGI, 共站同覆盖名, 物理站名, 小区名称, 使用频段, 是否覆盖层, 小区所属区域, 路测网格, 经度, 纬度, sectionid
        """
        df = pd.read_excel(excel_path, dtype_backend="numpy_nullable")
        
        # 标准化列名（按模板列名）
        column_mapping = {
            "CGI": "CGI",
            "共站同覆盖名": "共站同覆盖名",
            "物理站名": "物理站名",
            "小区名称": "小区名称",
            "使用频段": "使用频段",
            "是否覆盖层": "是否覆盖层",
            "小区所属区域": "小区所属区域",
            "路测网格": "路测网格",
            "经度": "经度",
            "纬度": "纬度",
            "sectionid": "sectionid",
        }
        
        # 重命名列（匹配模板）
        rename_cols = {}
        for col in df.columns:
            col_stripped = col.strip()
            if col_stripped in column_mapping:
                rename_cols[col] = column_mapping[col_stripped]
        
        df = df.rename(columns=rename_cols)
        
        # 确保必要列存在
        required_cols = ["CGI", "共站同覆盖名"]
        for col in required_cols:
            if col not in df.columns:
                raise ValueError(f"Excel缺少必要列: {col}")
        
        # 提取sectionid（从共站同覆盖名）
        if "sectionid" not in df.columns or df["sectionid"].isna().all():
            df["sectionid"] = df["共站同覆盖名"].apply(extract_sectionid)
        
        # 选择数据库中存在的列（按模板结构）
        db_cols = ["CGI", "共站同覆盖名", "物理站名", "小区名称", "使用频段", 
                   "是否覆盖层", "小区所属区域", "路测网格", "经度", "纬度", "sectionid"]
        existing_cols = [c for c in db_cols if c in df.columns]
        df = df[existing_cols]
        
        # 清空或追加
        if replace:
            self.conn.execute("DELETE FROM 共站同覆盖小区表")
        
        # 批量插入（使用UPSERT处理重复）
        self.conn.register("df_import", df)
        self.conn.execute("""
            INSERT OR REPLACE INTO 共站同覆盖小区表 
            SELECT CGI, 共站同覆盖名, 物理站名, 小区名称, 使用频段, 是否覆盖层, 
                   小区所属区域, 路测网格, 经度, 纬度, sectionid,
                   CURRENT_TIMESTAMP, CURRENT_TIMESTAMP
            FROM df_import
        """)
        self.conn.unregister("df_import")
        
        return len(df)
    
    def export_to_excel(self, excel_path: Path | str) -> int:
        """导出到Excel - 使用标准模板列顺序"""
        df = self.get_all()
        
        # 按标准模板顺序排列列
        template_cols = ["CGI", "共站同覆盖名", "物理站名", "小区名称", "使用频段", 
                        "是否覆盖层", "小区所属区域", "路测网格", "经度", "纬度", "sectionid"]
        
        # 只保留模板中存在的列
        available_cols = [c for c in template_cols if c in df.columns]
        # 添加其他列（如时间戳）
        other_cols = [c for c in df.columns if c not in template_cols]
        final_cols = available_cols + other_cols
        
        df = df[final_cols]
        df.to_excel(excel_path, index=False)
        return len(df)
    
    def get_count(self) -> int:
        """获取记录总数"""
        result = self.conn.execute("SELECT COUNT(*) FROM 共站同覆盖小区表").fetchone()
        return result[0] if result else 0
    
    def get_mapping_dict(self) -> dict[str, dict]:
        """获取CGI到记录的映射字典（供其他模块使用）"""
        df = self.get_all()
        result = {}
        for _, row in df.iterrows():
            cgi = str(row.get("CGI", ""))
            if cgi and cgi != "nan":
                coverage = row.get("是否覆盖层", row.get("覆盖层", ""))
                coverage = str(coverage) if pd.notna(coverage) else ""
                result[cgi] = {
                    "共站同覆盖名": row.get("共站同覆盖名", ""),
                    "sectionid": row.get("sectionid"),
                    "覆盖层": coverage,
                    "小区所属区域": row.get("小区所属区域", ""),
                    "路测网格": row.get("路测网格", ""),
                    "乡镇街道": row.get("乡镇街道", ""),
                    "是否覆盖层": coverage,
                }
        return result


# ==============================================================================
# 物理表模块：数据库操作 (DuckDB) - 保持向后兼容
# ==============================================================================

def connect_physical_db(base_dir: str):
    """连接 DuckDB 数据库（物理表专用）- 现在使用统一数据库"""
    return get_unified_db_connection()


def init_physical_database(conn):
    """初始化物理表数据库表结构 - 现在使用统一数据库"""
    # 确保统一库中的共站同覆盖表已就绪（持久化共享，不在此处重建/覆盖）
    init_unified_database(conn)
    
    # Drop existing tables to ensure fresh start (clear any leftover data)
    conn.execute("DROP TABLE IF EXISTS 物理表汇总")
    conn.execute("DROP TABLE IF EXISTS 原始小区表")
    conn.execute("""
        CREATE TABLE 原始小区表 (
            CGI TEXT PRIMARY KEY,
            网络制式 TEXT,
            小区名称 TEXT,
            物理站 TEXT,
            物理站ID TEXT,
            经度 DOUBLE,
            纬度 DOUBLE,
            方位角 DOUBLE,
            天线名 TEXT,
            挂高 DOUBLE,
            厂家 TEXT,
            BAND TEXT,
            BAND_A TEXT,
            频点 DOUBLE,
            站点类型 TEXT,
            网元状态 TEXT,
            覆盖类型 TEXT,
            乡镇街道 TEXT,
            一级标签 TEXT,
            路测网格 TEXT,
            来源文件 TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE 物理表汇总 (
            物理站 TEXT,
            物理站ID TEXT,
            CGI TEXT PRIMARY KEY,
            小区名称 TEXT,
            共站同覆盖名 TEXT,
            sectionid INTEGER,
            经度 DOUBLE,
            纬度 DOUBLE,
            网络制式 TEXT,
            BAND TEXT,
            BAND_A TEXT,
            频点 DOUBLE,
            路测网格 TEXT,
            区域 TEXT,
            督办网格ID TEXT,
            督办网格中文名 TEXT,
            乡镇 TEXT,
            物理扇区LTE制式 TEXT,
            物理站LTE制式 TEXT,
            物理扇区制式 TEXT,
            物理站制式 TEXT,
            覆盖层 TEXT,
            小区所属区域 TEXT,
            天线名 TEXT,
            方位角 DOUBLE,
            挂高 DOUBLE,
            厂家 TEXT,
            站点类型 TEXT,
            网元状态 TEXT,
            覆盖类型 TEXT,
            乡镇街道 TEXT,
            一级标签 TEXT,
            网络结构4G TEXT,
            共站制式情况 TEXT,
            物理站名_距离聚合 TEXT,
            物理站LTE制式_距离聚合 TEXT,
            物理站制式_距离聚合 TEXT,
            共站制式情况_距离聚合 TEXT
        )
    """)


FILE_PATTERNS = {
    "5g_week": "5G小区容量-周*.xlsx",
    "5g_day": "5G小区容量报表*.xlsx",
    "5g_mr": "5GMR覆盖-小区天*.xlsx",
    "5g_kpi": "5G小区性能KPI报表*.xlsx",
    "4g_week": "重要场景-周*.xlsx",
    "4g_day": "重要场景-天*.xlsx",
    "4g_mr": "4GMR覆盖-小区天*.xlsx",
    "cog_coverage": "共站同覆盖小区_4g_5g.xlsx",
}

# 物理表文件模式（使用 DATA_DIR）
PHYSICAL_FILE_PATTERNS = {
    "nr_cellant": "*_nr_*.xlsx",  # 5G工参
    "lte_cellant": "*_lte_*.xlsx",  # 4G工参
}

OUTPUT_5G = BASE_DIR / "合成_容量表_5G.xlsx"
OUTPUT_4G = BASE_DIR / "合成_容量表_4G.xlsx"
OUTPUT_45G = BASE_DIR / "容量表_45G.xlsx"
DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")
TIME_COLUMN_CANDIDATES = {
    "start": ["记录开始时间", "开始时间"],
    "end": ["记录结束时间", "结束时间"],
}

ProgressCallback = Callable[[int, str], None]
LogCallback = Callable[[str], None]


class SourceFileError(RuntimeError):
    pass


def setup_logging() -> logging.Logger:
    """配置日志系统，每天一个日志文件，自动清理超过7天的日志"""
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    
    # 清理过期日志
    cleanup_old_logs()
    
    # 创建日志记录器
    logger = logging.getLogger("CapPhysCombine")
    logger.setLevel(logging.DEBUG)
    
    # 避免重复添加处理器
    if logger.handlers:
        return logger
    
    # 格式化
    formatter = logging.Formatter(
        "%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    
    # 文件处理器 - 每天一个文件
    today = datetime.now().strftime("%Y%m%d")
    log_file = LOG_DIR / f"app_{today}.log"
    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)
    
    # 控制台处理器
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger


def cleanup_old_logs() -> None:
    """删除超过7天的日志文件"""
    if not LOG_DIR.exists():
        return
    
    cutoff_date = datetime.now() - timedelta(days=LOG_RETENTION_DAYS)
    deleted_count = 0
    
    for log_file in LOG_DIR.glob("app_*.log"):
        try:
            file_time = datetime.fromtimestamp(log_file.stat().st_mtime)
            if file_time < cutoff_date:
                log_file.unlink()
                deleted_count += 1
        except OSError:
            continue
    
    if deleted_count > 0:
        logging.getLogger("CapPhysCombine").info(f"已清理 {deleted_count} 个过期日志文件")


def get_logger() -> logging.Logger:
    """获取已配置的日志记录器"""
    logger = logging.getLogger("CapPhysCombine")
    if not logger.handlers:
        return setup_logging()
    return logger


class GuiLogger:
    def __init__(self, callback: LogCallback | None = None) -> None:
        self.callback = callback
        self._logger = get_logger()

    def log(self, message: str) -> None:
        self._logger.info(message)
        if self.callback:
            self.callback(message)


class GuiProgress:
    def __init__(self, callback: ProgressCallback | None = None, logger: GuiLogger | None = None) -> None:
        self.callback = callback
        self.logger = logger or GuiLogger()

    def update(self, value: int, message: str) -> None:
        value = max(0, min(100, value))
        if self.callback:
            self.callback(value, message)
        self.logger.log(message)


def pick_latest_file(pattern: str) -> Path:
    matches = sorted(DATA_DIR.glob(pattern))
    if not matches:
        raise SourceFileError(f"未找到匹配文件: {DATA_DIR / pattern}")

    def sort_key(path: Path) -> tuple[tuple[str, ...], str]:
        dates = tuple(DATE_RE.findall(path.name))
        return dates, path.name

    return max(matches, key=sort_key)


def import_cog_coverage_to_db(
    conn: duckdb.DuckDBPyConnection,
    logger: GuiLogger | None = None
) -> bool:
    """导入共站同覆盖小区表到数据库 - 现在使用统一数据库
    
    注意: 此函数保留以保持兼容性，实际上数据已保存在统一数据库中
    用户可以通过GUI管理共站同覆盖表
    """
    logger = logger or GuiLogger()
    logger.log("共站同覆盖表已从统一数据库加载")
    return True


def load_cog_coverage_mapping(
    conn: duckdb.DuckDBPyConnection,
    logger: GuiLogger | None = None
) -> pd.DataFrame:
    """从统一数据库加载共站同覆盖小区表，返回 CGI -> 共站同覆盖名 的映射表
    
    现在使用统一数据库的共站同覆盖表，两个功能共享同一个数据源
    """
    logger = logger or GuiLogger()
    
    # 使用统一数据库的共站同覆盖表
    try:
        with CogCoverageManager() as mgr:
            count = mgr.get_count()
            if count == 0:
                logger.log("统一数据库中共站同覆盖表为空，请通过管理界面导入")
                return pd.DataFrame(columns=["CGI", "共站同覆盖名"])
            
            logger.log(f"从统一数据库加载共站同覆盖映射表（{count} 条记录）")
            
            # 获取所有数据
            df = mgr.get_all()
            if df.empty:
                return pd.DataFrame(columns=["CGI", "共站同覆盖名"])
            
            # 选择需要的列
            required_cols = ["CGI", "共站同覆盖名"]
            extra_cols = ["路测网格", "乡镇街道", "是否覆盖层", "小区所属区域", "覆盖层"]
            
            available_cols = [c for c in required_cols + extra_cols if c in df.columns]
            mapping = df[available_cols].copy()
            
            # 确保CGI为字符串类型
            mapping["CGI"] = mapping["CGI"].astype(str)
            
            # 去重（保留第一个）
            mapping = mapping.drop_duplicates(subset=["CGI"], keep="first")
            
            extra_loaded = [c for c in extra_cols if c in mapping.columns]
            logger.log(f"共站同覆盖映射表加载完成，共 {len(mapping)} 条映射，额外字段: {extra_loaded}")
            return mapping
            
    except Exception as e:
        logger.log(f"从统一数据库加载共站同覆盖表失败: {e}")
        return pd.DataFrame(columns=["CGI", "共站同覆盖名"])


def read_excel(path: Path) -> pd.DataFrame:
    """快速读取 Excel 文件"""
    return pd.read_excel(
        path,
        engine="openpyxl",
        dtype_backend="numpy_nullable",
    )


DB_PATH = BASE_DIR / "capphys.db"
CHUNK_SIZE = 5000


def get_db_connection() -> duckdb.DuckDBPyConnection:
    return duckdb.connect(str(DB_PATH))


def init_db() -> None:
    if DB_PATH.exists():
        try:
            DB_PATH.unlink()
        except OSError:
            pass


def excel_to_db(
    path: Path,
    table_name: str,
    conn: duckdb.DuckDBPyConnection,
    logger: GuiLogger | None = None,
    chunk_size: int = CHUNK_SIZE,
    append: bool = False,
) -> int:
    logger = logger or GuiLogger()
    mode = "追加" if append else "导入"
    logger.log(f"  分批{mode} {path.name} -> {table_name} (每批 {chunk_size} 行)")
    
    from openpyxl import load_workbook
    
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    
    try:
        headers = next(rows_iter)
    except StopIteration:
        wb.close()
        return 0
    
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(headers)]
    total_rows = 0
    chunk: list[tuple] = []
    first_write = not append
    
    for row in rows_iter:
        chunk.append(row)
        if len(chunk) >= chunk_size:
            df_chunk = pd.DataFrame(chunk, columns=headers)
            tmp_name = f"_tmp_{table_name}"
            conn.register(tmp_name, df_chunk)
            if first_write:
                conn.execute(f"CREATE OR REPLACE TABLE \"{table_name}\" AS SELECT * FROM {tmp_name}")
                first_write = False
            else:
                conn.execute(f"INSERT INTO \"{table_name}\" SELECT * FROM {tmp_name}")
            total_rows += len(chunk)
            chunk = []
    
    if chunk:
        df_chunk = pd.DataFrame(chunk, columns=headers)
        tmp_name = f"_tmp_{table_name}"
        conn.register(tmp_name, df_chunk)
        if first_write:
            conn.execute(f"CREATE OR REPLACE TABLE \"{table_name}\" AS SELECT * FROM {tmp_name}")
        else:
            conn.execute(f"INSERT INTO \"{table_name}\" SELECT * FROM {tmp_name}")
        total_rows += len(chunk)
    
    wb.close()
    return total_rows


def small_excel_to_db(
    path: Path,
    table_name: str,
    conn: duckdb.DuckDBPyConnection,
    logger: GuiLogger | None = None,
    append: bool = False,
) -> int:
    logger = logger or GuiLogger()
    df = pd.read_excel(path, engine="openpyxl", dtype_backend="numpy_nullable")
    df.columns = [str(c) for c in df.columns]
    
    mode = "追加" if append else "导入"
    tmp_name = f"_tmp_{table_name}"
    conn.register(tmp_name, df)
    if append:
        conn.execute(f"INSERT INTO \"{table_name}\" SELECT * FROM {tmp_name}")
    else:
        conn.execute(f"CREATE OR REPLACE TABLE \"{table_name}\" AS SELECT * FROM {tmp_name}")
    
    logger.log(f"  {mode} {path.name} -> {table_name} ({len(df)} 行)")
    return len(df)


def db_to_dataframe(query: str, conn: duckdb.DuckDBPyConnection) -> pd.DataFrame:
    return conn.execute(query).fetchdf()


def table_exists(conn: duckdb.DuckDBPyConnection, table_name: str) -> bool:
    result = conn.execute(f"SELECT 1 FROM information_schema.tables WHERE table_name = '{table_name}'").fetchone()
    return result is not None


def get_table_columns(conn: duckdb.DuckDBPyConnection, table_name: str) -> list[str]:
    result = conn.execute(f"DESCRIBE \"{table_name}\"").fetchdf()
    return list(result["column_name"])


def first_existing(df: pd.DataFrame, columns: Iterable[str], default=None):
    for column in columns:
        if column in df.columns:
            return df[column]
    return pd.Series(default, index=df.index)


def to_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def safe_divide(numerator: pd.Series, denominator: pd.Series) -> pd.Series:
    denominator = denominator.replace(0, pd.NA)
    return numerator.div(denominator)


def normalize_datetime(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce")


def build_output_paths(timestamp: str | None = None) -> dict[str, Path]:
    suffix = f"_{timestamp}" if timestamp else ""
    return {
        "5g": BASE_DIR / f"合成_容量表_5G{suffix}.xlsx",
        "4g": BASE_DIR / f"合成_容量表_4G{suffix}.xlsx",
        "45g": BASE_DIR / f"容量表_45G{suffix}.xlsx",
    }


def first_valid_timestamp(df: pd.DataFrame, candidates: list[str]) -> pd.Timestamp | None:
    for column in candidates:
        if column not in df.columns:
            continue
        series = pd.to_datetime(df[column], errors="coerce").dropna()
        if not series.empty:
            return series.iloc[0]
    return None


def pick_matching_files(pattern: str) -> list[Path]:
    matches = list(DATA_DIR.glob(pattern))
    if not matches:
        return []
    
    def sort_key(path: Path) -> tuple[tuple[str, ...], str]:
        dates = tuple(DATE_RE.findall(path.name))
        return dates, path.name
    
    return sorted(matches, key=sort_key)


def resolve_output_timestamp(conn: duckdb.DuckDBPyConnection) -> str | None:
    for source_name in ["5g_week", "4g_week"]:
        if not table_exists(conn, source_name):
            continue
        df = db_to_dataframe(f"SELECT * FROM \"{source_name}\" LIMIT 100", conn)
        if df.empty:
            continue
        start_time = first_valid_timestamp(df, TIME_COLUMN_CANDIDATES["start"])
        end_time = first_valid_timestamp(df, TIME_COLUMN_CANDIDATES["end"])
        if start_time is not None and end_time is not None:
            return f"{start_time:%Y%m%d}_{end_time:%Y%m%d}"
    return None


LARGE_TABLES = {"5g_day", "5g_mr", "5g_kpi", "4g_day", "4g_mr"}


def load_sources_to_db(conn: duckdb.DuckDBPyConnection, logger: GuiLogger | None = None) -> dict[str, list[Path]]:
    """将源数据文件导入数据库（单线程顺序导入，避免DuckDB并发问题）"""
    logger = logger or GuiLogger()
    
    selected: dict[str, list[Path]] = {}
    
    for name, pattern in FILE_PATTERNS.items():
        files = pick_matching_files(pattern)
        if files:
            selected[name] = files
    
    logger.log("使用以下源文件：")
    for name, files in selected.items():
        if len(files) == 1:
            logger.log(f"- {name}: {files[0].name}")
        else:
            logger.log(f"- {name}: {len(files)} 个文件")
            for f in files:
                logger.log(f"    - {f.name}")
    
    logger.log("开始导入数据...")
    start_total = time.perf_counter()
    
    # 顺序导入，避免DuckDB并发问题
    for name, files in selected.items():
        try:
            start = time.perf_counter()
            total_rows = 0
            for i, path in enumerate(files):
                append = i > 0
                if name in LARGE_TABLES:
                    rows = excel_to_db(path, name, conn, logger, append=append)
                else:
                    rows = small_excel_to_db(path, name, conn, logger, append=append)
                total_rows += rows
            
            col_count = len(get_table_columns(conn, name))
            elapsed = time.perf_counter() - start
            logger.log(f"导入完成 [{name}]: {total_rows} 行 x {col_count} 列 (耗时 {elapsed:.1f}s)")
        except Exception as e:
            logger.log(f"导入失败 [{name}]: {e}")
    
    # 导入共站同覆盖表（如果存在）
    import_cog_coverage_to_db(conn, logger)
    
    elapsed_total = time.perf_counter() - start_total
    logger.log(f"所有数据导入完成，总耗时 {elapsed_total:.1f}s")
    
    return selected


def load_small_table(conn: duckdb.DuckDBPyConnection, name: str) -> pd.DataFrame:
    if not table_exists(conn, name):
        return pd.DataFrame()
    return db_to_dataframe(f"SELECT * FROM \"{name}\"", conn)


def apply_sector_mapping(table: pd.DataFrame, mapping: pd.DataFrame, cgi_column: str, logger: GuiLogger | None = None) -> pd.DataFrame:
    """用共站同覆盖映射表更新容量表的扇区字段"""
    logger = logger or GuiLogger()
    
    if mapping.empty or cgi_column not in table.columns:
        logger.log(f"跳过扇区更新（CGI列: {cgi_column}）")
        return table
    
    table = table.copy()
    table[cgi_column] = table[cgi_column].astype(str)
    
    # 需要映射的字段列表
    mapping_cols = ["共站同覆盖名", "路测网格", "乡镇街道", "是否覆盖层", "小区所属区域"]
    mapping_cols = [col for col in mapping_cols if col in mapping.columns]
    
    # 遍历映射每个字段
    for col in mapping_cols:
        # 创建映射字典
        col_mapping = dict(zip(mapping["CGI"], mapping[col]))
        target_col = "扇区" if col == "共站同覆盖名" else col
        if target_col not in table.columns:
            table[target_col] = pd.NA
        original_count = table[target_col].notna().sum()
        table[target_col] = table[cgi_column].map(col_mapping)
        updated_count = table[target_col].notna().sum()
        logger.log(f"字段 [{target_col}] 更新完成: 匹配 {updated_count} 条（新增 {updated_count - original_count} 条）")
    
    return table


def build_5g_table(conn: duckdb.DuckDBPyConnection, logger: GuiLogger | None = None) -> pd.DataFrame:
    logger = logger or GuiLogger()
    
    logger.log("  [5G] 在数据库中进行日表聚合...")
    
    day_cols = get_table_columns(conn, "5g_day")
    util_col_5g = "忙时小区PRB利用率" if "忙时小区PRB利用率" in day_cols else "忙时小区PRB利用率(%)"
    
    conn.execute("DROP TABLE IF EXISTS _5g_day_agg")
    conn.execute("DROP TABLE IF EXISTS _5g_day_weekday")
    conn.execute("DROP TABLE IF EXISTS _5g_day_weekend")
    
    conn.execute(f"""
        CREATE TABLE _5g_day_agg AS
        SELECT
            CAST(NCGI AS VARCHAR) AS NCGI,
            AVG("{util_col_5g}") AS 自忙时利用率,
            AVG("日RLC层上下行总流量(G)") AS 日均流量,
            AVG("忙时上行PRB平均利用率(%)") AS 自忙时上行PRB平均利用率,
            AVG("忙时下行PRB平均利用率(%)") AS 自忙时下行PRB平均利用率,
            AVG("忙时PDCCH信道CCE占用率(%)") AS 自忙时PDCCH信道CCE占用率,
            AVG("RRC连接最大数-忙时") AS 自忙时RRC连接最大数,
            AVG("RRC连接平均数-忙时") AS 自忙时有效RRC连接平均数,
            AVG("忙时RLC层上行业务字节数(G)") AS 自忙时上行流量,
            AVG("忙时RLC层下行业务字节数(G)") AS 自忙时下行流量,
            AVG("忙时RLC层上行业务字节数(G)") + AVG("忙时RLC层下行业务字节数(G)") AS 自忙时总流量,
            AVG("RRC连接最大数-忙时") AS 自忙时有效RRC连接最大数
        FROM "5g_day"
        GROUP BY CAST(NCGI AS VARCHAR)
    """)
    
    conn.execute(f"""
        CREATE TABLE _5g_day_weekday AS
        SELECT
            CAST(NCGI AS VARCHAR) AS NCGI,
            AVG("{util_col_5g}") AS 工作日自忙时利用率,
            AVG("日RLC层上下行总流量(G)") AS 工作日日均流量,
            AVG("RRC连接最大数-忙时") AS 工作日自忙时RRC连接最大数
        FROM "5g_day"
        WHERE CAST(extract(dayofweek FROM CAST("记录开始时间" AS TIMESTAMP)) AS INTEGER) IN (2, 3, 4, 5, 6)
        GROUP BY CAST(NCGI AS VARCHAR)
    """)
    
    conn.execute(f"""
        CREATE TABLE _5g_day_weekend AS
        SELECT
            CAST(NCGI AS VARCHAR) AS NCGI,
            AVG("{util_col_5g}") AS 周末自忙时利用率,
            AVG("日RLC层上下行总流量(G)") AS 周末日均流量,
            AVG("RRC连接最大数-忙时") AS 周末自忙时RRC连接最大数
        FROM "5g_day"
        WHERE CAST(extract(dayofweek FROM CAST("记录开始时间" AS TIMESTAMP)) AS INTEGER) IN (1, 7)
        GROUP BY CAST(NCGI AS VARCHAR)
    """)

    conn.execute("DROP TABLE IF EXISTS _5g_day_zero_stats")
    conn.execute("""
        CREATE TABLE _5g_day_zero_stats AS
        WITH day_raw AS (
            SELECT
                CAST(NCGI AS VARCHAR) AS NCGI,
                CAST("记录开始时间" AS TIMESTAMP) AS 记录开始时间,
                COALESCE("日RLC层上下行总流量(G)", 0) AS 日流量
            FROM "5g_day"
        ),
        zero_stats AS (
            SELECT
                NCGI,
                SUM(CASE WHEN CAST(extract(dayofweek FROM 记录开始时间) AS INTEGER) IN (2, 3, 4, 5, 6)
                         AND 日流量 = 0 THEN 1 ELSE 0 END) AS 工作日零流量天数,
                SUM(CASE WHEN CAST(extract(dayofweek FROM 记录开始时间) AS INTEGER) IN (1, 7)
                         AND 日流量 = 0 THEN 1 ELSE 0 END) AS 周末零流量天数
            FROM day_raw
            GROUP BY NCGI
        ),
        top3 AS (
            SELECT
                NCGI,
                AVG(日流量) AS 最大3天流量均值
            FROM (
                SELECT
                    NCGI,
                    日流量,
                    ROW_NUMBER() OVER (PARTITION BY NCGI ORDER BY 日流量 DESC) AS rn
                FROM day_raw
            ) ranked
            WHERE rn <= 3
            GROUP BY NCGI
        )
        SELECT
            z.NCGI,
            z.工作日零流量天数,
            z.周末零流量天数,
            t.最大3天流量均值
        FROM zero_stats z
        LEFT JOIN top3 t ON z.NCGI = t.NCGI
    """)
    
    logger.log("  [5G] 在数据库中进行 MR 表聚合...")
    conn.execute("DROP TABLE IF EXISTS _5g_mr_agg")
    conn.execute("""
        CREATE TABLE _5g_mr_agg AS
        SELECT
            CAST("小区NCGI" AS VARCHAR) AS NCGI,
            AVG("移动RSRP采样的总采样点") AS MRO移动总采样点,
            SUM("移动RSRP采样强于-110采样点") AS MRO强于110采样点合计,
            SUM("移动RSRP采样的总采样点") AS MRO总采样点合计,
            AVG("移动平均TA(M)") AS 平均TA米,
            CASE 
                WHEN SUM("移动RSRP采样的总采样点") = 0 OR SUM("移动RSRP采样的总采样点") IS NULL THEN NULL 
                ELSE SUM("移动RSRP采样强于-110采样点") * 1.0 / SUM("移动RSRP采样的总采样点") 
            END AS MRO移动覆盖率
        FROM "5g_mr"
        GROUP BY CAST("小区NCGI" AS VARCHAR)
    """)
    
    logger.log("  [5G] 在数据库中进行 KPI 表聚合...")
    conn.execute("DROP TABLE IF EXISTS _5g_kpi_agg")
    conn.execute("""
        CREATE TABLE _5g_kpi_agg AS
        SELECT
            CAST(NCGI AS VARCHAR) AS NCGI,
            AVG("VoNR语音话务量") AS VoNR语音话务量
        FROM "5g_kpi"
        GROUP BY CAST(NCGI AS VARCHAR)
    """)
    
    logger.log("  [5G] 加载周表并去重...")
    week_df = load_small_table(conn, "5g_week")
    if week_df.empty:
        return pd.DataFrame()
    
    week_df["NCGI"] = week_df["NCGI"].astype(str)
    week_unique = week_df.drop_duplicates(subset=["NCGI"]).copy()
    
    logger.log("  [5G] 读取聚合结果并合并...")
    day_agg = db_to_dataframe("SELECT * FROM _5g_day_agg", conn)
    weekday_agg = db_to_dataframe("SELECT * FROM _5g_day_weekday", conn)
    weekend_agg = db_to_dataframe("SELECT * FROM _5g_day_weekend", conn)
    zero_stats = db_to_dataframe("SELECT * FROM _5g_day_zero_stats", conn)
    mr_agg = db_to_dataframe("SELECT * FROM _5g_mr_agg", conn)
    kpi_agg = db_to_dataframe("SELECT * FROM _5g_kpi_agg", conn)
    
    result = week_unique.merge(day_agg, on="NCGI", how="left")
    result = result.merge(weekday_agg, on="NCGI", how="left")
    result = result.merge(weekend_agg, on="NCGI", how="left")
    result = result.merge(zero_stats, on="NCGI", how="left")
    result = result.merge(mr_agg, on="NCGI", how="left")
    result = result.merge(kpi_agg, on="NCGI", how="left")
    
    logger.log("  [5G] 计算流量系数、长尾分类等派生字段...")
    
    avg_traffic = result["日均流量"].mean(skipna=True)
    result["流量系数"] = result["日均流量"] / avg_traffic if pd.notna(avg_traffic) and avg_traffic != 0 else pd.NA
    tail_threshold = result["日均流量"].quantile(0.3)
    result["流量排名升序"] = result["日均流量"].rank(method="min", ascending=True)
    
    is_na_traffic = result["日均流量"].isna()
    is_tail = (result["日均流量"] <= tail_threshold) & ~is_na_traffic
    is_zero = result["日均流量"] == 0
    is_high_util = result["自忙时利用率"].notna() & (result["自忙时利用率"] > 20)
    
    result["长尾小区"] = pd.NA
    result.loc[is_tail & is_zero, "长尾小区"] = "长尾具体原因待确认"
    result.loc[is_tail & ~is_zero & is_high_util, "长尾小区"] = "长尾待观察"
    result.loc[is_tail & ~is_zero & ~is_high_util, "长尾小区"] = "长尾需处理"
    
    result["流量是否正常"] = pd.NA
    result.loc[result["流量系数"] < 0.2, "流量是否正常"] = "低流量系数小区"
    result.loc[(result["流量系数"] >= 0.2) & (result["流量系数"] < 3), "流量是否正常"] = "正常"
    result.loc[result["流量系数"] >= 3, "流量是否正常"] = "高流量系数小区"
    
    result["负荷情况"] = pd.NA
    result.loc[result["自忙时利用率"].notna() & (result["自忙时利用率"] > 80), "负荷情况"] = "负荷高小区"
    result.loc[result["自忙时利用率"].isna() | (result["自忙时利用率"] <= 80), "负荷情况"] = "正常"
    
    result["记录开始时间"] = first_existing(result, ["记录开始时间"])
    result["记录结束时间"] = first_existing(result, ["记录结束时间"])
    result["地市"] = first_existing(result, ["地市"])
    result["网元状态"] = first_existing(result, ["网元状态"])
    result["小区名称"] = first_existing(result, ["小区名称"])
    result["band"] = first_existing(result, ["使用频段"])
    result["覆盖类型"] = first_existing(result, ["覆盖类型"])
    result["场景 V容量表"] = first_existing(result, ["场景1", "一级场景"])
    result["TYPE"] = pd.NA
    result["是否全省高负荷预警小区（集团口径）"] = pd.NA
    result["是否高负荷待扩容小区"] = first_existing(result, ["是否高负荷待扩容小区", "是否高负荷"])
    result["是否全省高负荷预警小区（省内口径）"] = pd.NA
    result["物理站"] = first_existing(result, ["站点名称"])
    
    ordered_columns = [
        "记录开始时间", "记录结束时间", "地市", "NCGI", "网元状态", "小区名称", "扇区", "band",
        "覆盖类型", "场景 V容量表", "TYPE", "流量是否正常", "负荷情况", "流量排名升序", "长尾小区",
        "自忙时利用率", "日均流量", "VoNR语音话务量", "MRO移动总采样点", "MRO移动覆盖率", "平均TA米",
        "工作日自忙时利用率", "工作日日均流量", "工作日自忙时RRC连接最大数",
        "周末自忙时利用率", "周末日均流量", "周末自忙时RRC连接最大数",
        "工作日零流量天数", "周末零流量天数", "最大3天流量均值",
        "自忙时上行PRB平均利用率", "自忙时下行PRB平均利用率", "自忙时PDCCH信道CCE占用率",
        "自忙时有效RRC连接最大数", "自忙时RRC连接最大数", "自忙时有效RRC连接平均数",
        "自忙时总流量", "自忙时上行流量", "自忙时下行流量",
        "是否全省高负荷预警小区（集团口径）", "是否高负荷待扩容小区", "是否全省高负荷预警小区（省内口径）",
        "流量系数", "物理站",
    ]
    return result.reindex(columns=ordered_columns)


def build_4g_table(conn: duckdb.DuckDBPyConnection, logger: GuiLogger | None = None) -> pd.DataFrame:
    logger = logger or GuiLogger()
    
    logger.log("  [4G] 在数据库中进行日表聚合...")
    
    conn.execute("DROP TABLE IF EXISTS _4g_day_temp")
    conn.execute("DROP TABLE IF EXISTS _4g_day_agg")
    conn.execute("DROP TABLE IF EXISTS _4g_day_weekday")
    conn.execute("DROP TABLE IF EXISTS _4g_day_weekend")
    
    conn.execute("""
        CREATE TABLE _4g_day_temp AS
        SELECT
            CAST(CGI AS VARCHAR) AS CGI,
            CASE WHEN "自忙时上行PRB平均利用率" > "自忙时下行PRB平均利用率" THEN "自忙时上行PRB平均利用率" ELSE "自忙时下行PRB平均利用率" END AS 自忙时利用率,
            "日4G流量（GB）",
            "自忙时上行PRB平均利用率",
            "自忙时下行PRB平均利用率",
            "自忙时PDCCH信道CCE占用率",
            "自忙时有效RRC连接最大数",
            "自忙时RRC连接最大数",
            "自忙时有效RRC连接平均数",
            "自忙时空口上行业务字节数",
            "自忙时空口下行业务字节数",
            "记录开始时间"
        FROM "4g_day"
    """)
    
    conn.execute("""
        CREATE TABLE _4g_day_agg AS
        SELECT
            CGI,
            AVG(自忙时利用率) AS 自忙时利用率,
            AVG("日4G流量（GB）") AS 日均流量,
            AVG("自忙时上行PRB平均利用率") AS 自忙时上行PRB平均利用率,
            AVG("自忙时下行PRB平均利用率") AS 自忙时下行PRB平均利用率,
            AVG("自忙时PDCCH信道CCE占用率") AS 自忙时PDCCH信道CCE占用率,
            AVG("自忙时有效RRC连接最大数") AS 自忙时有效RRC连接最大数,
            AVG("自忙时RRC连接最大数") AS 自忙时RRC连接最大数,
            AVG("自忙时有效RRC连接平均数") AS 自忙时有效RRC连接平均数,
            AVG("自忙时空口上行业务字节数") AS 自忙时上行流量,
            AVG("自忙时空口下行业务字节数") AS 自忙时下行流量,
            AVG("自忙时空口上行业务字节数") + AVG("自忙时空口下行业务字节数") AS 自忙时总流量
        FROM _4g_day_temp
        GROUP BY CGI
    """)
    
    conn.execute("""
        CREATE TABLE _4g_day_weekday AS
        SELECT
            CGI,
            AVG(自忙时利用率) AS 工作日自忙时利用率,
            AVG("日4G流量（GB）") AS 工作日日均流量,
            AVG("自忙时RRC连接最大数") AS 工作日自忙时RRC连接最大数
        FROM _4g_day_temp
        WHERE CAST(extract(dayofweek FROM CAST("记录开始时间" AS TIMESTAMP)) AS INTEGER) IN (2, 3, 4, 5, 6)
        GROUP BY CGI
    """)
    
    conn.execute("""
        CREATE TABLE _4g_day_weekend AS
        SELECT
            CGI,
            AVG(自忙时利用率) AS 周末自忙时利用率,
            AVG("日4G流量（GB）") AS 周末日均流量,
            AVG("自忙时RRC连接最大数") AS 周末自忙时RRC连接最大数
        FROM _4g_day_temp
        WHERE CAST(extract(dayofweek FROM CAST("记录开始时间" AS TIMESTAMP)) AS INTEGER) IN (1, 7)
        GROUP BY CGI
    """)
    
    logger.log("  [4G] 在数据库中进行 MR 表聚合...")
    conn.execute("DROP TABLE IF EXISTS _4g_mr_agg")
    conn.execute("""
        CREATE TABLE _4g_mr_agg AS
        SELECT
            CAST(cgi AS VARCHAR) AS CGI,
            AVG("MRO移动总采样点") AS MRO移动总采样点,
            SUM("MRO移动大于等于负110DBM的采样点数") AS MRO有效点合计,
            SUM("MRO移动总采样点") AS MRO总采样点合计,
            AVG("平均TA") AS 平均TA米,
            CASE 
                WHEN SUM("MRO移动总采样点") = 0 OR SUM("MRO移动总采样点") IS NULL THEN NULL 
                ELSE SUM("MRO移动大于等于负110DBM的采样点数") * 1.0 / SUM("MRO移动总采样点") 
            END AS MRO移动覆盖率
        FROM "4g_mr"
        GROUP BY CAST(cgi AS VARCHAR)
    """)
    
    logger.log("  [4G] 在数据库中进行周表指标聚合...")
    conn.execute("DROP TABLE IF EXISTS _4g_week_metrics")
    conn.execute("""
        CREATE TABLE _4g_week_metrics AS
        SELECT
            CAST(CGI AS VARCHAR) AS CGI,
            AVG("自忙时上行PRB平均利用率") AS week_上行PRB,
            AVG("自忙时下行PRB平均利用率") AS week_下行PRB,
            AVG("自忙时PDCCH信道CCE占用率") AS week_PDCCH
        FROM "4g_week"
        GROUP BY CAST(CGI AS VARCHAR)
    """)
    
    logger.log("  [4G] 加载周表并去重...")
    week_df = load_small_table(conn, "4g_week")
    if week_df.empty:
        return pd.DataFrame()
    
    week_df["CGI"] = week_df["CGI"].astype(str)
    week_unique = week_df.drop_duplicates(subset=["CGI"]).copy()
    
    logger.log("  [4G] 读取聚合结果并合并...")
    day_agg = db_to_dataframe("SELECT * FROM _4g_day_agg", conn)
    weekday_agg = db_to_dataframe("SELECT * FROM _4g_day_weekday", conn)
    weekend_agg = db_to_dataframe("SELECT * FROM _4g_day_weekend", conn)
    mr_agg = db_to_dataframe("SELECT * FROM _4g_mr_agg", conn)
    week_metrics = db_to_dataframe("SELECT * FROM _4g_week_metrics", conn)
    
    result = week_unique.merge(day_agg, on="CGI", how="left")
    result = result.merge(week_metrics, on="CGI", how="left")
    result = result.merge(weekday_agg, on="CGI", how="left")
    result = result.merge(weekend_agg, on="CGI", how="left")
    result = result.merge(mr_agg, on="CGI", how="left")
    
    logger.log("  [4G] 计算流量系数、长尾分类等派生字段...")
    
    avg_traffic = result["日均流量"].mean(skipna=True)
    result["流量系数"] = result["日均流量"] / avg_traffic if pd.notna(avg_traffic) and avg_traffic != 0 else pd.NA
    result["流量排名升序"] = result["日均流量"].rank(method="min", ascending=True)
    tail_threshold = result["日均流量"].quantile(0.3)
    
    result["流量是否正常"] = pd.NA
    result.loc[result["流量系数"] < 0.2, "流量是否正常"] = "低流量系数小区"
    result.loc[(result["流量系数"] >= 0.2) & (result["流量系数"] < 3), "流量是否正常"] = "正常"
    result.loc[result["流量系数"] >= 3, "流量是否正常"] = "高流量系数小区"
    
    name = result["小区名称"].fillna("").astype(str)
    is_rdc_dc = name.str.contains("RDC|DC-|RGS|GS-", regex=True, na=False)
    is_rd = name.str.contains("RD-", regex=True, na=False)
    
    util = result["自忙时利用率"]
    result["负荷情况"] = "正常"
    result.loc[is_rdc_dc & (util > 90), "负荷情况"] = "负荷高小区"
    result.loc[is_rd & (util > 70) & ~is_rdc_dc, "负荷情况"] = "负荷高小区"
    result.loc[~is_rdc_dc & ~is_rd & (util > 50), "负荷情况"] = "负荷高小区"
    result.loc[util.isna(), "负荷情况"] = "正常"
    
    is_na_traffic = result["日均流量"].isna()
    is_tail = (result["日均流量"] <= tail_threshold) & ~is_na_traffic
    is_zero = result["日均流量"] == 0
    is_high_util = result["自忙时利用率"].notna() & (result["自忙时利用率"] > 20)
    
    result["长尾小区"] = pd.NA
    result.loc[is_tail & is_zero, "长尾小区"] = "具体原因待确认"
    result.loc[is_tail & ~is_zero & is_high_util, "长尾小区"] = "长尾待观察"
    result.loc[is_tail & ~is_zero & ~is_high_util, "长尾小区"] = "长尾需处理"
    
    result["记录开始时间"] = first_existing(result, ["记录开始时间"])
    result["记录结束时间"] = first_existing(result, ["记录结束时间"])
    result["地市"] = first_existing(result, ["所属地市"])
    result["网元状态"] = first_existing(result, ["网元状态"])
    result["小区名称"] = first_existing(result, ["小区名称"])
    result["band"] = first_existing(result, ["使用频段", "频点"])
    result["场景 V容量表"] = first_existing(result, ["场景"])
    result["TYPE"] = pd.NA
    result["语音话务量Erl （VOLTE/VoNR）"] = first_existing(result, ["VOLTE语音话务量"])
    result["是否全省高负荷预警小区（集团口径）"] = first_existing(result, ["是否高流量预警小区"])
    result["是否高负荷待扩容小区"] = first_existing(result, ["是否高负荷待扩容小区"])
    result["是否全省高负荷预警小区（省内口径）"] = first_existing(result, ["是否高流量预警小区"])
    result["物理站"] = first_existing(result, ["所属站点名称"])
    
    if "week_上行PRB" in result.columns:
        if "自忙时上行PRB平均利用率_y" in result.columns:
            result["自忙时上行PRB平均利用率"] = result["自忙时上行PRB平均利用率_y"].combine_first(result["week_上行PRB"])
        else:
            result["自忙时上行PRB平均利用率"] = result["自忙时上行PRB平均利用率"].combine_first(result["week_上行PRB"])
    
    ordered_columns = [
        "记录开始时间", "记录结束时间", "地市", "CGI", "网元状态", "小区名称", "扇区", "band",
        "场景 V容量表", "TYPE", "流量是否正常", "负荷情况", "流量排名升序", "长尾小区",
        "自忙时利用率", "日均流量", "语音话务量Erl （VOLTE/VoNR）", "MRO移动总采样点", "MRO移动覆盖率", "平均TA米",
        "工作日自忙时利用率", "工作日日均流量", "工作日自忙时RRC连接最大数",
        "周末自忙时利用率", "周末日均流量", "周末自忙时RRC连接最大数",
        "自忙时上行PRB平均利用率", "自忙时下行PRB平均利用率", "自忙时PDCCH信道CCE占用率",
        "自忙时有效RRC连接最大数", "自忙时RRC连接最大数", "自忙时有效RRC连接平均数",
        "自忙时总流量", "自忙时上行流量", "自忙时下行流量",
        "是否全省高负荷预警小区（集团口径）", "是否高负荷待扩容小区", "是否全省高负荷预警小区（省内口径）",
        "流量系数", "物理站",
    ]
    return result.reindex(columns=ordered_columns)


def build_45g_table(table_5g: pd.DataFrame, table_4g: pd.DataFrame) -> pd.DataFrame:
    merged_5g = table_5g.copy()
    merged_4g = table_4g.copy()

    merged_5g.insert(0, "网络制式", "5G")
    merged_4g.insert(0, "网络制式", "4G")

    merged_5g = merged_5g.rename(columns={"NCGI": "CGI/NCGI", "VoNR语音话务量": "语音话务量Erl （VOLTE/VoNR）"})
    merged_4g = merged_4g.rename(columns={"CGI": "CGI/NCGI"})

    all_columns = ["网络制式"] + [column for column in merged_5g.columns if column != "网络制式"]
    for column in all_columns:
        if column not in merged_5g.columns:
            merged_5g[column] = pd.NA
        if column not in merged_4g.columns:
            merged_4g[column] = pd.NA

    merged = pd.concat(
        [merged_5g[all_columns], merged_4g[all_columns]],
        ignore_index=True,
        sort=False,
    )
    return merged


def _safe_div(numerator, denominator):
    if denominator in (0, 0.0, None) or pd.isna(denominator):
        return pd.NA
    return numerator / denominator


CAPACITY_BAND_TO_STANDARD = {
    "D频段": "D",
    "F频段": "F",
    "E频段": "E",
    "A频段": "A",
    "5G-3DMIMO": "5G-3Dmimo",
    "5G-3Dmimo": "5G-3Dmimo",
    "3D-MIMO": "5G-3Dmimo",
}


def _normalize_capacity_band(value) -> str:
    band = normalize_text(value)
    if not band:
        return ""
    return CAPACITY_BAND_TO_STANDARD.get(band, band)


def _build_loweff_band_lookups(
    table_4g: pd.DataFrame, table_5g: pd.DataFrame
) -> tuple[dict, dict, dict, dict]:
    """汇总同扇区/物理站的纯4G频段与含5G频段字符串。"""
    frames: list[pd.DataFrame] = []
    for table in (table_4g, table_5g):
        if table is None or table.empty:
            continue
        n = len(table)
        phys = (
            table["物理站"].map(normalize_text)
            if "物理站" in table.columns
            else pd.Series([""] * n, index=table.index)
        )
        sector = (
            table["扇区"].map(normalize_text)
            if "扇区" in table.columns
            else pd.Series([""] * n, index=table.index)
        )
        band = (
            table["band"].map(_normalize_capacity_band)
            if "band" in table.columns
            else pd.Series([""] * n, index=table.index)
        )
        frames.append(pd.DataFrame({"物理站": phys, "扇区": sector, "BAND": band}))
    if not frames:
        return {}, {}, {}, {}
    work = pd.concat(frames, ignore_index=True)
    work = work[work["BAND"].astype(bool)]
    sector_all: dict = {}
    sector_lte: dict = {}
    station_all: dict = {}
    station_lte: dict = {}
    sector_work = work[work["扇区"].astype(bool)]
    if not sector_work.empty:
        sector_all, sector_lte = build_band_aggregations(sector_work, "扇区", LTE_BANDS)
    station_work = work[work["物理站"].astype(bool)]
    if not station_work.empty:
        station_all, station_lte = build_band_aggregations(station_work, "物理站", LTE_BANDS)
    return sector_lte, sector_all, station_lte, station_all


FULL_4G_EVAL_COLUMNS = [
    "网络制式",
    "CGI/NCGI",
    "小区名称",
    "物理站",
    "扇区",
    "同扇区纯4G频段",
    "同扇区包含5G频段",
    "物理站纯4G频段",
    "物理站包含5G频段",
    "地市",
    "忙时利用率",
    "忙时流量",
    "自忙时有效RRC连接平均数",
    "扇区等效利用率_拆除前",
    "扇区等效单载波流量_拆除前",
    "小区拆除后扇区等效利用率（<40%）",
    "扇区等效单载波流量_拆除后",
    "能否减容",
]

LTE_BAND_M_VALUES = {
    "3D-MIMO": 2.5,
    "5G-3Dmimo": 2.5,
    "5G-3DMIMO": 2.5,
    "FDD1800": 1.5,
    "E": 1.0,
    "E频段": 1.0,
    "D": 1.0,
    "D频段": 1.0,
    "F": 1.0,
    "F频段": 1.0,
    "F1": 1.0,
    "A": 0.75,
    "A频段": 0.75,
    "F2": 0.5,
    "FDD900": 0.75,
}


def _resolve_lte_band_m(row) -> float | None:
    raw = normalize_text(row.get("band")) or normalize_text(row.get("BAND_A"))
    if not raw:
        return None
    if raw in LTE_BAND_M_VALUES:
        return float(LTE_BAND_M_VALUES[raw])
    std = _normalize_capacity_band(raw)
    if std in LTE_BAND_M_VALUES:
        return float(LTE_BAND_M_VALUES[std])
    return None


def build_low_efficiency_table(
    table_5g: pd.DataFrame, table_4g: pd.DataFrame, table_45g: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rows_5g: list[dict[str, object]] = []
    rows_4g: list[dict[str, object]] = []
    rows_4g_full: list[dict[str, object]] = []
    sector_lte_map, sector_all_map, station_lte_map, station_all_map = _build_loweff_band_lookups(
        table_4g, table_5g
    )

    nr_eval_bands = {"2.6GHz", "4.9GHz"}
    if not table_5g.empty:
        for _, row in table_5g.iterrows():
            band = normalize_text(row.get("band"))
            if band not in nr_eval_bands:
                continue

            max3_avg = pd.to_numeric(row.get("最大3天流量均值"), errors="coerce")
            work_zero_days = pd.to_numeric(row.get("工作日零流量天数"), errors="coerce")
            weekend_zero_days = pd.to_numeric(row.get("周末零流量天数"), errors="coerce")
            coverage = normalize_text(row.get("覆盖类型"))

            # 5G零效益：工作日零流量>=2 且 周末零流量>=1
            zero_trigger = (
                pd.notna(work_zero_days)
                and work_zero_days >= 2
                and pd.notna(weekend_zero_days)
                and weekend_zero_days >= 1
            )

            # 5G低效益：最大3天流量均值，宏站/室外<3GB，室分/室内<2GB
            if coverage in {"室内", "室分"}:
                site_kind = "室分"
                low_threshold = 2.0
            else:
                site_kind = "宏站"
                low_threshold = 3.0
            low_trigger = pd.notna(max3_avg) and max3_avg < low_threshold

            if not (zero_trigger or low_trigger):
                continue

            if zero_trigger:
                low_type = "零效益小区"
                low_reason = (
                    f"工作日零流量{int(work_zero_days)}天+周末零流量{int(weekend_zero_days)}天"
                )
            else:
                low_type = "低效益小区"
                low_reason = (
                    f"{site_kind}最大3天流量均值<{low_threshold:g}GB（{max3_avg:.2f}GB）"
                )

            rows_5g.append({
                "网络制式": "5G",
                "CGI/NCGI": row.get("NCGI"),
                "小区名称": row.get("小区名称"),
                "物理站": row.get("物理站"),
                "扇区": row.get("扇区"),
                "地市": row.get("地市"),
                "band": band,
                "覆盖类型": coverage or pd.NA,
                "站点类型判定": site_kind,
                "忙时利用率": row.get("自忙时利用率"),
                "忙时流量": row.get("日均流量"),
                "自忙时有效RRC连接平均数": row.get("自忙时有效RRC连接平均数"),
                "工作日零流量天数": work_zero_days,
                "周末零流量天数": weekend_zero_days,
                "最大3天流量均值": max3_avg,
                "低效类型": low_type,
                "低效原因": low_reason,
                "优先级": 1 if zero_trigger else 2,
            })

    if not table_4g.empty:
        for sector, group in table_4g.groupby(table_4g["扇区"].fillna("").astype(str), dropna=False):
            sector_df = group.copy()
            has_sector = bool(sector)

            denom = 0.0
            ps = 0.0
            traffic_sum = 0.0
            if has_sector:
                for _, srow in sector_df.iterrows():
                    m = _resolve_lte_band_m(srow)
                    if m is None:
                        continue
                    util = pd.to_numeric(srow.get("自忙时利用率"), errors="coerce")
                    traffic = pd.to_numeric(srow.get("日均流量"), errors="coerce")
                    if pd.notna(util):
                        ps += m * util
                    if pd.notna(traffic):
                        traffic_sum += traffic
                    denom += m

            before_util = round(float(ps / denom), 2) if has_sector and denom > 0 else pd.NA
            before_traffic = round(float(traffic_sum / denom), 2) if has_sector and denom > 0 else pd.NA
            sector_candidates: list[dict[str, object]] = []

            for _, row in sector_df.iterrows():
                station = normalize_text(row.get("物理站"))
                this_m = _resolve_lte_band_m(row)
                after_util = pd.NA
                after_traffic = pd.NA
                low_reason = ""
                impact = None
                if has_sector and this_m is not None and denom > this_m:
                    new_denom = denom - this_m
                    if new_denom > 0:
                        after_util = round(float(ps / new_denom), 2)
                        after_traffic = round(float(traffic_sum / new_denom), 2)
                        if after_util < 40 and after_traffic < 20:
                            low_reason = "拆除后扇区等效利用率<40% 且 拆除后扇区等效单载波流量<20GB"
                            impact = float(after_util) + float(after_traffic)

                base_row = {
                    "网络制式": "4G",
                    "CGI/NCGI": row.get("CGI"),
                    "小区名称": row.get("小区名称"),
                    "物理站": row.get("物理站"),
                    "扇区": row.get("扇区"),
                    "同扇区纯4G频段": sector_lte_map.get(sector, "") if has_sector else "",
                    "同扇区包含5G频段": sector_all_map.get(sector, "") if has_sector else "",
                    "物理站纯4G频段": station_lte_map.get(station, ""),
                    "物理站包含5G频段": station_all_map.get(station, ""),
                    "地市": row.get("地市"),
                    "忙时利用率": row.get("自忙时利用率"),
                    "忙时流量": row.get("日均流量"),
                    "自忙时有效RRC连接平均数": row.get("自忙时有效RRC连接平均数"),
                    "扇区等效利用率_拆除前": before_util,
                    "扇区等效单载波流量_拆除前": before_traffic,
                    "小区拆除后扇区等效利用率（<40%）": after_util,
                    "扇区等效单载波流量_拆除后": after_traffic,
                    "能否减容": "是" if low_reason else "否",
                    "低效原因": low_reason,
                }
                rows_4g_full.append(base_row)

                if low_reason and impact is not None:
                    candidate = dict(base_row)
                    candidate["优先级"] = 1
                    candidate["影响值"] = impact
                    sector_candidates.append(candidate)

            if sector_candidates:
                best = sorted(
                    sector_candidates,
                    key=lambda x: (
                        x["影响值"],
                        x["小区拆除后扇区等效利用率（<40%）"],
                        x["扇区等效单载波流量_拆除后"],
                    ),
                )[0]
                best.pop("影响值", None)
                rows_4g.append(best)

    df5 = pd.DataFrame(rows_5g)
    if not df5.empty:
        df5 = df5.sort_values(by=["优先级", "低效类型", "最大3天流量均值"], ascending=[True, True, True])

    df4 = pd.DataFrame(rows_4g)
    if not df4.empty:
        df4 = df4.sort_values(
            by=["优先级", "小区拆除后扇区等效利用率（<40%）", "扇区等效单载波流量_拆除后"],
            ascending=[True, True, True],
        )

    df4_full = pd.DataFrame(rows_4g_full, columns=FULL_4G_EVAL_COLUMNS)
    if not df4_full.empty:
        df4_full = df4_full.reindex(columns=FULL_4G_EVAL_COLUMNS)

    nr_denom = 0
    if not table_5g.empty and "band" in table_5g.columns:
        bands = table_5g["band"].map(normalize_text)
        nr_denom = int(bands.isin({"2.6GHz", "4.9GHz"}).sum())
    zero_cnt = int((df5["低效类型"] == "零效益小区").sum()) if not df5.empty else 0
    low_cnt = int((df5["低效类型"] == "低效益小区").sum()) if not df5.empty else 0
    ratio = round(len(df5) / nr_denom, 6) if nr_denom > 0 else pd.NA

    summary_rows = [
        {"指标": "45G总数", "数值": len(table_45g)},
        {"指标": "2.6G/4.9G 5G小区总数", "数值": nr_denom},
        {"指标": "5G零效益数", "数值": zero_cnt},
        {"指标": "5G低效益数", "数值": low_cnt},
        {"指标": "5G低效数", "数值": len(df5)},
        {"指标": "5G低效占比", "数值": ratio},
        {"指标": "4G低效数", "数值": len(df4)},
        {"指标": "全量4G评估数", "数值": len(df4_full)},
        {"指标": "低效总数", "数值": len(df5) + len(df4)},
    ]
    summary = pd.DataFrame(summary_rows)
    return df5, df4, df4_full, summary

LOWEFF_OUTPUT_PATH = BASE_DIR / "低效小区结果.xlsx"


def _build_capacity_tables(conn, logger, progress: GuiProgress | None = None):
    if progress:
        progress.update(22, "数据导入完成，使用数据库进行聚合计算")

    timestamp = resolve_output_timestamp(conn)
    output_paths = build_output_paths(timestamp)
    if timestamp:
        logger.log(f"本次输出时间戳: {timestamp}")
    else:
        logger.log("未从周表中识别到开始/结束时间，输出文件将使用默认文件名")

    logger.log("开始生成 5G 容量表")
    start = time.perf_counter()
    table_5g = build_5g_table(conn, logger)
    elapsed = time.perf_counter() - start
    if progress:
        progress.update(45, f"5G表生成完成，共 {len(table_5g)} 条")
    logger.log(f"5G 容量表生成完成 (耗时 {elapsed:.1f}s)")

    cog_mapping = load_cog_coverage_mapping(conn, logger)
    if not cog_mapping.empty:
        table_5g = apply_sector_mapping(table_5g, cog_mapping, "NCGI", logger)

    logger.log("开始生成 4G 容量表")
    start = time.perf_counter()
    table_4g = build_4g_table(conn, logger)
    elapsed = time.perf_counter() - start
    if progress:
        progress.update(68, f"4G表生成完成，共 {len(table_4g)} 条")
    logger.log(f"4G 容量表生成完成 (耗时 {elapsed:.1f}s)")

    if not cog_mapping.empty:
        table_4g = apply_sector_mapping(table_4g, cog_mapping, "CGI", logger)

    return table_5g, table_4g, output_paths


def run_pipeline(
    progress_callback: ProgressCallback | None = None,
    log_callback: LogCallback | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    started_at = time.perf_counter()
    logger = GuiLogger(log_callback)
    logger.log("=" * 50)
    logger.log("容量表合成工具启动")
    logger.log("=" * 50)
    progress = GuiProgress(progress_callback, logger)
    logger.log("初始化数据库（低内存模式：使用 DuckDB 磁盘数据库）")
    init_db()
    conn = get_db_connection()
    try:
        progress.update(3, "开始导入数据到数据库")
        logger.log("准备扫描并导入源文件到 DuckDB 数据库")
        load_sources_to_db(conn, logger)
        table_5g, table_4g, output_paths = _build_capacity_tables(conn, logger, progress)
        logger.log("开始合并 45G 总表")
        start = time.perf_counter()
        table_45g = build_45g_table(table_5g, table_4g)
        elapsed = time.perf_counter() - start
        progress.update(80, f"45G总表生成完成，共 {len(table_45g)} 条")
        logger.log(f"45G 总表生成完成 (耗时 {elapsed:.1f}s)")
        logger.log("开始生成低效小区结果")
        loweff_5g, loweff_4g, loweff_4g_full, loweff_summary = build_low_efficiency_table(
            table_5g, table_4g, table_45g
        )
        logger.log(
            f"低效小区结果生成完成，5G {len(loweff_5g)} 条，4G {len(loweff_4g)} 条，"
            f"全量4G评估 {len(loweff_4g_full)} 条"
        )
        logger.log(f"开始写出文件: {output_paths['5g'].name}")
        start = time.perf_counter(); table_5g.to_excel(output_paths["5g"], index=False); elapsed = time.perf_counter() - start
        progress.update(88, f"已生成: {output_paths['5g'].name}"); logger.log(f"写出 {output_paths['5g'].name} 完成 (耗时 {elapsed:.1f}s)")
        logger.log(f"开始写出文件: {output_paths['4g'].name}")
        start = time.perf_counter(); table_4g.to_excel(output_paths["4g"], index=False); elapsed = time.perf_counter() - start
        progress.update(94, f"已生成: {output_paths['4g'].name}"); logger.log(f"写出 {output_paths['4g'].name} 完成 (耗时 {elapsed:.1f}s)")
        logger.log(f"开始写出文件: {output_paths['45g'].name}")
        start = time.perf_counter(); table_45g.to_excel(output_paths["45g"], index=False); elapsed = time.perf_counter() - start
        progress.update(96, f"已生成: {output_paths['45g'].name}"); logger.log(f"写出 {output_paths['45g'].name} 完成 (耗时 {elapsed:.1f}s)")
        with pd.ExcelWriter(LOWEFF_OUTPUT_PATH) as writer:
            loweff_5g.to_excel(writer, index=False, sheet_name="5G低效明细")
            loweff_4g.to_excel(writer, index=False, sheet_name="4G低效明细")
            loweff_4g_full.to_excel(writer, index=False, sheet_name="全量4G小区评估")
            loweff_summary.to_excel(writer, index=False, sheet_name="统计汇总")
        logger.log(f"写出 {LOWEFF_OUTPUT_PATH.name} 完成")
        progress.update(100, f"已生成: {LOWEFF_OUTPUT_PATH.name}")
        elapsed_seconds = time.perf_counter() - started_at
        logger.log(f"5G表记录数: {len(table_5g)}")
        logger.log(f"4G表记录数: {len(table_4g)}")
        logger.log(f"45G总表记录数: {len(table_45g)}")
        logger.log(f"总耗时: {elapsed_seconds:.2f} 秒")
        logger.log("全部处理完成")
        logger.log("=" * 50)
        get_logger().info(f"处理完成，耗时 {elapsed_seconds:.2f} 秒")
        return table_5g, table_4g, table_45g
    finally:
        conn.close()
        try:
            if DB_PATH.exists():
                DB_PATH.unlink()
                logger.log("临时数据库已清理")
        except OSError:
            pass


def main() -> None:
    run_pipeline()


# ==============================================================================
# 物理表模块：扇区工具 (sector_tools.py)
# ==============================================================================

SECTION_PATTERN = re.compile(r"(?:扇区|S)(\d+)", re.IGNORECASE)
BAND_SECTION_HINTS = {
    "F1": 1,
    "F2": 2,
    "E1": 1,
    "E2": 2,
    "E3": 3,
    "D1": 1,
    "D3": 3,
    "D7": 7,
    "D8": 8,
}


def normalize_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def extract_section_no_from_name(name):
    text = normalize_text(name)
    if not text:
        return None
    m = SECTION_PATTERN.search(text)
    if m:
        return int(m.group(1))
    return None


def _guess_section_no(row):
    name_no = extract_section_no_from_name(row.get("小区名称", ""))
    if name_no is not None:
        return name_no
    band_a = normalize_text(row.get("BAND_A", ""))
    if band_a in BAND_SECTION_HINTS:
        return BAND_SECTION_HINTS[band_a]
    return None


def detect_sector_conflicts(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    required = {"物理站", "BAND", "sectionid", "CGI", "小区名称", "共站同覆盖名"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"缺少列: {', '.join(sorted(missing))}")

    work = df.copy()
    work["物理站"] = work["物理站"].map(normalize_text)
    work["BAND"] = work["BAND"].map(normalize_text)
    work["sectionid"] = pd.to_numeric(work["sectionid"], errors="coerce")

    conflict_frames = [grp for _, grp in work.groupby(["物理站", "BAND", "sectionid"], dropna=False) if len(grp) > 1]
    if not conflict_frames:
        return work.iloc[0:0].copy()
    return pd.concat(conflict_frames, ignore_index=True)


def suggest_sector_fixes(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    work = df.copy()
    work["建议扇区号"] = work.apply(_guess_section_no, axis=1)
    return work


def recompute_derived_fields(df: pd.DataFrame) -> pd.DataFrame:
    """修正 sectionid 后，重算依赖列，保持汇总口径一致。"""
    work = df.copy()
    if "物理扇区制式" in work.columns:
        valid_sector = work[work["共站同覆盖名"].notna() & (work["共站同覆盖名"] != "")]
        if len(valid_sector) > 0:
            sector_all, sector_lte = build_band_aggregations(valid_sector, "共站同覆盖名", LTE_BANDS)
            work["物理扇区制式"] = work["共站同覆盖名"].map(sector_all).fillna("")
            work["物理扇区LTE制式"] = work["共站同覆盖名"].map(sector_lte).fillna("")
    if "物理站制式" in work.columns:
        valid_station = work[work["物理站"].notna() & (work["物理站"] != "")]
        if len(valid_station) > 0:
            station_all, station_lte = build_band_aggregations(valid_station, "物理站", LTE_BANDS)
            work["物理站制式"] = work["物理站"].map(station_all).fillna("")
            work["物理站LTE制式"] = work["物理站"].map(station_lte).fillna("")
    if {"物理扇区LTE制式", "覆盖层"}.issubset(work.columns):
        work = apply_lte_network_structure(work)
    if "物理站制式" in work.columns:
        work["共站制式情况"] = work["物理站制式"].apply(co_site_coverage_type)
    return work


def auto_fix_sector_conflicts(df: pd.DataFrame):
    if df.empty:
        return df.copy(), df.copy(), df.copy()
    work = df.copy()
    conflict_rows = []
    fix_rows = []

    group_cols = ["物理站", "BAND", "sectionid"]
    for _, grp in work.groupby(group_cols, dropna=False):
        if len(grp) <= 1:
            continue

        grp = grp.copy()
        grp["建议扇区号"] = grp.apply(_guess_section_no, axis=1)
        used = set()

        for idx, row in grp.sort_values(by=["建议扇区号", "方位角"], na_position="last").iterrows():
            conflict_rows.append(row)
            suggested = row.get("建议扇区号")
            new_section = None
            if pd.notna(suggested):
                suggested = int(suggested)
                if suggested not in used:
                    new_section = suggested
                    used.add(suggested)
            if new_section is None:
                candidate = 1
                while candidate in used:
                    candidate += 1
                new_section = candidate
                used.add(candidate)

            old_section = row.get("sectionid")
            old_name = row.get("共站同覆盖名", "")
            base_name = normalize_text(old_name)
            if base_name:
                base_name = re.sub(r"(?:-?扇区\d+|-?S\d+)$", "", base_name)
            else:
                base_name = normalize_text(row.get("物理站", ""))

            new_name = f"{base_name}-扇区{new_section}"
            work.at[idx, "sectionid"] = new_section
            work.at[idx, "共站同覆盖名"] = new_name
            fix_rows.append(
                {
                    "CGI": row.get("CGI", ""),
                    "小区名称": row.get("小区名称", ""),
                    "物理站": row.get("物理站", ""),
                    "站点类型": row.get("站点类型", ""),
                    "BAND": row.get("BAND", ""),
                    "原sectionid": old_section,
                    "新sectionid": new_section,
                    "原共站同覆盖名": old_name,
                    "新共站同覆盖名": new_name,
                    "建议扇区号": row.get("建议扇区号"),
                }
            )

    conflict_df = pd.DataFrame(conflict_rows).drop_duplicates()
    fix_df = pd.DataFrame(fix_rows)
    work = recompute_derived_fields(work)
    return work, conflict_df, fix_df


# ==============================================================================
# 物理表模块：聚合器主类 (aggregator.py)
# ==============================================================================

def build_cc_lookup(cc_df: pd.DataFrame) -> dict:
    lookup = {}
    if cc_df.empty:
        return lookup
    for _, row in cc_df.iterrows():
        cgi = str(row.get("CGI", ""))
        if cgi and cgi != "nan":
            lookup[cgi] = {
                "共站同覆盖名": row.get("共站同覆盖名", ""),
                "sectionid": row.get("sectionid"),
                "覆盖层": str(row.get("覆盖层", "")) if pd.notna(row.get("覆盖层")) else "",
                "小区所属区域": row.get("小区所属区域", ""),
                "路测网格": row.get("路测网格", ""),
            }
    return lookup


def merge_cc_and_spatial_fields(cgis, cc_lookup, loadtest_grid_ids, cell_loadtest_grids):
    common_names, sectionids, coverage_layers, region_from_cc, loadtest_grids = (
        [],
        [],
        [],
        [],
        [],
    )
    for i, cgi in enumerate(cgis):
        cc_info = cc_lookup.get(cgi, {})
        common_names.append(cc_info.get("共站同覆盖名", ""))
        sectionids.append(cc_info.get("sectionid"))
        coverage_layers.append(cc_info.get("覆盖层", ""))
        region_from_cc.append(cc_info.get("小区所属区域", ""))
        cc_grid = cc_info.get("路测网格", "")
        spatial_grid = loadtest_grid_ids[i]
        cell_grid = cell_loadtest_grids[i] if i < len(cell_loadtest_grids) else ""
        final_grid = (
            cc_grid
            if cc_grid and str(cc_grid) != "nan"
            else (spatial_grid if spatial_grid else cell_grid)
        )
        loadtest_grids.append(final_grid)
    return common_names, sectionids, coverage_layers, region_from_cc, loadtest_grids


class PhysicalTableAggregator:
    """单地市物理表汇总（编排层）。"""

    def __init__(self, base_dir: str):
        self.base_dir = base_dir
        self.conn = None

    def init_database(self):
        self.conn = connect_physical_db(self.base_dir)
        init_physical_database(self.conn)

    def aggregate_physical_table(self, data_dir: Path | None = None) -> pd.DataFrame:
        """物理表汇总 - 从 DATA_DIR 读取工参文件
        
        Args:
            data_dir: 数据目录，默认为 DATA_DIR
        """
        if data_dir is None:
            data_dir = DATA_DIR
        
        self.init_database()

        print("=" * 50)
        print("开始物理表汇总...")
        print("=" * 50)
        print(f"数据目录: {data_dir}")

        # 使用模式匹配查找工参文件（类似容量表的方式）
        nr_files = sorted(data_dir.glob("*_nr_*.xlsx"))
        lte_files = sorted(data_dir.glob("*_lte_*.xlsx"))
        
        # 过滤临时文件
        nr_files = [f for f in nr_files if not f.name.startswith(".~")]
        lte_files = [f for f in lte_files if not f.name.startswith(".~")]

        print(f"\n找到 {len(nr_files)} 个5G工参文件, {len(lte_files)} 个4G工参文件")

        frames = []
        for nr_file in nr_files:
            df = read_nr_cellant(nr_file)
            frames.append(df)
            print(f"已读取5G工参: {nr_file.name}, {len(df)} 条")
        for lte_file in lte_files:
            df = read_lte_cellant(lte_file)
            frames.append(df)
            print(f"已读取4G工参: {lte_file.name}, {len(df)} 条")

        all_cells = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        print(f"\n共读取 {len(all_cells)} 条小区数据")
        
        if all_cells.empty:
            print("警告: 未读取到任何小区数据，请检查数据目录")
            return all_cells
        
        # 按CGI去重，保留第一个出现的记录
        before_dedup = len(all_cells)
        all_cells = all_cells.drop_duplicates(subset=["CGI"], keep="first")
        after_dedup = len(all_cells)
        if before_dedup != after_dedup:
            print(f"去重后剩余 {after_dedup} 条记录（去除 {before_dedup - after_dedup} 条重复CGI）")
            
        # 使用 DuckDB 方式写入数据 - 直接注册 DataFrame 为临时表
        self.conn.register("df_cells", all_cells)
        self.conn.execute("""
            INSERT INTO 原始小区表 
            SELECT 
                CAST(CGI AS VARCHAR),
                CAST(网络制式 AS VARCHAR),
                CAST(小区名称 AS VARCHAR),
                CAST(物理站 AS VARCHAR),
                CAST(物理站ID AS VARCHAR),
                CAST(经度 AS DOUBLE),
                CAST(纬度 AS DOUBLE),
                CAST(方位角 AS DOUBLE),
                CAST(天线名 AS VARCHAR),
                CAST(挂高 AS DOUBLE),
                CAST(厂家 AS VARCHAR),
                CAST(BAND AS VARCHAR),
                CAST(BAND_A AS VARCHAR),
                CAST(频点 AS DOUBLE),
                CAST(站点类型 AS VARCHAR),
                CAST(网元状态 AS VARCHAR),
                CAST(覆盖类型 AS VARCHAR),
                CAST(乡镇街道 AS VARCHAR),
                CAST(一级标签 AS VARCHAR),
                CAST(路测网格 AS VARCHAR),
                CAST(来源文件 AS VARCHAR)
            FROM df_cells
        """)
        self.conn.unregister("df_cells")

        # 从统一数据库加载共站同覆盖表（两个功能共享同一个数据源）
        print("\n从统一数据库加载共站同覆盖小区表...")
        cc_df = pd.DataFrame()
        cc_lookup = {}
        
        try:
            with CogCoverageManager() as mgr:
                cc_count = mgr.get_count()
                if cc_count > 0:
                    cc_df = mgr.get_all()
                    cc_lookup = mgr.get_mapping_dict()
                    print(f"已从统一数据库加载共站同覆盖小区表: {cc_count} 条记录")
                else:
                    print("警告: 统一数据库中共站同覆盖表为空，请通过管理界面导入")
                    
                    # 尝试从Excel文件加载作为备用
                    cc_files = sorted(data_dir.glob("共站同覆盖小区*.xlsx"))
                    cc_files = [f for f in cc_files if not f.name.startswith(".~")]
                    
                    if cc_files:
                        cc_path = cc_files[0]
                        cc_df = read_common_coverage(cc_path)
                        cc_lookup = build_cc_lookup(cc_df)
                        print(f"已从Excel加载共站同覆盖小区表: {cc_path.name}, {len(cc_lookup)} 条")
                        print("提示: 建议导入到统一数据库以便两个功能共享")
        except Exception as e:
            print(f"从统一数据库加载失败: {e}")
            print("尝试从Excel文件加载...")
            cc_files = sorted(data_dir.glob("共站同覆盖小区*.xlsx"))
            cc_files = [f for f in cc_files if not f.name.startswith(".~")]
            
            if cc_files:
                cc_path = cc_files[0]
                cc_df = read_common_coverage(cc_path)
                cc_lookup = build_cc_lookup(cc_df)
                print(f"已从Excel加载共站同覆盖小区表: {cc_path.name}, {len(cc_lookup)} 条")

        # 加载地理边界文件（从BASE_DIR子目录）
        print("\n加载地理边界文件（带空间索引）...")
        base = Path(self.base_dir)
        loadtest_gdf, loadtest_sindex = load_geojson_with_index(
            base / "路测网格" / "loadtest_grid.geojson"
        )
        region_gdf, region_sindex = load_geojson_with_index(
            base / "区域" / "阳江五区域.geojson"
        )
        grid_gdf, grid_sindex = load_geojson_with_index(
            base / "网格" / "grid_yj.geojson"
        )
        town_gdf, town_sindex = load_geojson_with_index(
            base / "乡镇" / "镇界.geojson"
        )

        print(f"共站同覆盖小区表索引构建完成: {len(cc_lookup)} 条CGI记录")

        print("\n执行批量空间查询...")
        lons = all_cells["经度"].values
        lats = all_cells["纬度"].values
        loadtest_results = get_grid_by_coords_batch(
            loadtest_gdf, loadtest_sindex, lons, lats
        )
        region_results = get_grid_by_coords_batch(region_gdf, region_sindex, lons, lats)
        grid_results = get_grid_by_coords_batch(grid_gdf, grid_sindex, lons, lats)
        town_results = get_grid_by_coords_batch(town_gdf, town_sindex, lons, lats)

        print("提取地理信息...")
        loadtest_grid_ids = [
            r.get("grid_id", r.get("id", "")) if r is not None else None
            for r in loadtest_results
        ]
        region_names = [
            r.get("name", r.get("区域名称", "")) if r is not None else None
            for r in region_results
        ]
        grid_ids = [r.get("序号", None) if r is not None else None for r in grid_results]
        grid_names = [r.get("中文名", None) if r is not None else None for r in grid_results]
        town_names = [
            r.get("name", r.get("镇区名称", None)) if r is not None else None
            for r in town_results
        ]

        print("构建汇总表...")
        cgis = all_cells["CGI"].astype(str).values
        cell_loadtest_grids = all_cells["路测网格"].values
        (
            common_names,
            sectionids,
            coverage_layers,
            region_from_cc,
            loadtest_grids,
        ) = merge_cc_and_spatial_fields(
            cgis, cc_lookup, loadtest_grid_ids, cell_loadtest_grids
        )

        agg_df = pd.DataFrame(
            {
                "物理站": all_cells["物理站"].values,
                "物理站ID": all_cells["物理站ID"].values,
                "CGI": cgis,
                "小区名称": all_cells["小区名称"].values,
                "共站同覆盖名": common_names,
                "sectionid": sectionids,
                "经度": lons,
                "纬度": lats,
                "网络制式": all_cells["网络制式"].values,
                "BAND": all_cells["BAND"].values,
                "BAND_A": all_cells["BAND_A"].values,
                "频点": all_cells["频点"].values,
                "路测网格": loadtest_grids,
                "区域": [r if r else region_from_cc[i] for i, r in enumerate(region_names)],
                "督办网格ID": grid_ids,
                "督办网格中文名": grid_names,
                "乡镇": town_names,
                "物理扇区LTE制式": "",
                "物理站LTE制式": "",
                "物理扇区制式": "",
                "物理站制式": "",
                "覆盖层": coverage_layers,
                "小区所属区域": region_from_cc,
                "天线名": all_cells["天线名"].values,
                "方位角": all_cells["方位角"].values,
                "挂高": all_cells["挂高"].values,
                "厂家": all_cells["厂家"].values,
                "站点类型": all_cells["站点类型"].values,
                "网元状态": all_cells["网元状态"].values,
                "覆盖类型": all_cells["覆盖类型"].values,
                "乡镇街道": all_cells["乡镇街道"].values,
                "一级标签": all_cells["一级标签"].values,
                "网络结构4G": "",
                "共站制式情况": "",
            }
        )

        print("\n计算制式汇总...")
        valid_sector = agg_df[
            agg_df["共站同覆盖名"].notna() & (agg_df["共站同覆盖名"] != "")
        ]
        if len(valid_sector) > 0:
            sector_all, sector_lte = build_band_aggregations(
                valid_sector, "共站同覆盖名", LTE_BANDS
            )
            agg_df["物理扇区制式"] = agg_df["共站同覆盖名"].map(sector_all).fillna("")
            agg_df["物理扇区LTE制式"] = (
                agg_df["共站同覆盖名"].map(sector_lte).fillna("")
            )

        valid_station = agg_df[agg_df["物理站"].notna() & (agg_df["物理站"] != "")]
        if len(valid_station) > 0:
            station_all, station_lte = build_band_aggregations(
                valid_station, "物理站", LTE_BANDS
            )
            agg_df["物理站制式"] = agg_df["物理站"].map(station_all).fillna("")
            agg_df["物理站LTE制式"] = agg_df["物理站"].map(station_lte).fillna("")

        print("计算4G网络结构...")
        agg_df = apply_lte_network_structure(agg_df)

        print("计算共站制式情况...")
        agg_df["共站制式情况"] = agg_df["物理站制式"].apply(co_site_coverage_type)

        agg_df = aggregate_by_distance(agg_df)
        # #endregion

        self.conn.register("df_agg", agg_df)
        self.conn.execute("""
            INSERT INTO 物理表汇总
            SELECT 
                CAST(物理站 AS VARCHAR),
                CAST(物理站ID AS VARCHAR),
                CAST(CGI AS VARCHAR),
                CAST(小区名称 AS VARCHAR),
                CAST(共站同覆盖名 AS VARCHAR),
                CAST(sectionid AS INTEGER),
                CAST(经度 AS DOUBLE),
                CAST(纬度 AS DOUBLE),
                CAST(网络制式 AS VARCHAR),
                CAST(BAND AS VARCHAR),
                CAST(BAND_A AS VARCHAR),
                CAST(频点 AS DOUBLE),
                CAST(路测网格 AS VARCHAR),
                CAST(区域 AS VARCHAR),
                CAST(督办网格ID AS VARCHAR),
                CAST(督办网格中文名 AS VARCHAR),
                CAST(乡镇 AS VARCHAR),
                CAST(物理扇区LTE制式 AS VARCHAR),
                CAST(物理站LTE制式 AS VARCHAR),
                CAST(物理扇区制式 AS VARCHAR),
                CAST(物理站制式 AS VARCHAR),
                CAST(覆盖层 AS VARCHAR),
                CAST(小区所属区域 AS VARCHAR),
                CAST(天线名 AS VARCHAR),
                CAST(方位角 AS DOUBLE),
                CAST(挂高 AS DOUBLE),
                CAST(厂家 AS VARCHAR),
                CAST(站点类型 AS VARCHAR),
                CAST(网元状态 AS VARCHAR),
                CAST(覆盖类型 AS VARCHAR),
                CAST(乡镇街道 AS VARCHAR),
                CAST(一级标签 AS VARCHAR),
                CAST(网络结构4G AS VARCHAR),
                CAST(共站制式情况 AS VARCHAR),
                CAST(物理站名_距离聚合 AS VARCHAR),
                CAST(物理站LTE制式_距离聚合 AS VARCHAR),
                CAST(物理站制式_距离聚合 AS VARCHAR),
                CAST(共站制式情况_距离聚合 AS VARCHAR)
            FROM df_agg
        """)
        self.conn.unregister("df_agg")

        print("\n" + "=" * 50)
        print("物理表汇总完成！")
        print("=" * 50)
        print(f"共处理 {len(agg_df)} 条记录")
        print(f"有共站同覆盖名的记录: {(agg_df['共站同覆盖名'] != '').sum()}")
        print(f"有sectionid的记录: {agg_df['sectionid'].notna().sum()}")
        print(f"有路测网格的记录: {(agg_df['路测网格'] != '').sum()}")
        print(f"有区域的记录: {(agg_df['区域'] != '').sum()}")
        print(f"有督办网格的记录: {agg_df['督办网格ID'].notna().sum()}")
        return agg_df

    def export_to_excel(self, output_path=None, agg_df=None):
        if output_path is None:
            output_path = os.path.join(self.base_dir, "物理表汇总结果.xlsx")
        if agg_df is None:
            agg_df = self.conn.execute("SELECT * FROM 物理表汇总").fetchdf()
        agg_df.to_excel(output_path, index=False)
        print(f"结果已导出到: {output_path}")

    def close(self):
        if self.conn:
            self.conn.close()


# ==================== 物理表功能 ====================

def run_physical_table_pipeline(
    base_dir: str,
    output_path: str | None = None,
    progress_callback: ProgressCallback | None = None,
    log_callback: LogCallback | None = None,
) -> pd.DataFrame | None:
    """运行物理表汇总流程
    
    Args:
        base_dir: 数据根目录，包含 45G工参、路测网格、区域、网格、乡镇 等子目录
        output_path: 输出Excel路径，默认为 base_dir/物理表汇总结果.xlsx
        progress_callback: 进度回调函数
        log_callback: 日志回调函数
        
    Returns:
        汇总后的DataFrame，失败返回None
    """
    logger = GuiLogger(log_callback)
    progress = GuiProgress(progress_callback, logger)
    
    if not PHYSICAL_TABLE_AVAILABLE:
        logger.log("错误: 物理表模块未安装，无法运行物理表功能")
        return None
    
    logger.log("=" * 50)
    logger.log("物理表汇总流程启动")
    logger.log("=" * 50)
    logger.log(f"数据目录: {base_dir}")
    
    if output_path is None:
        output_path = os.path.join(base_dir, "物理表汇总结果.xlsx")
    logger.log(f"输出文件: {output_path}")
    
    progress.update(5, "初始化物理表汇总器...")
    
    start_time = time.time()
    aggregator = PhysicalTableAggregator(base_dir)
    
    try:
        progress.update(20, "读取并汇总数据...")
        # 传入 DATA_DIR 作为数据来源
        agg_df = aggregator.aggregate_physical_table(DATA_DIR)
        
        progress.update(80, "导出Excel...")
        aggregator.export_to_excel(output_path, agg_df=agg_df)
        
        elapsed = time.time() - start_time
        progress.update(100, f"物理表汇总完成，共 {len(agg_df)} 条记录")
        logger.log(f"\n总耗时: {elapsed:.2f} 秒")
        logger.log(f"共处理 {len(agg_df)} 条记录")
        logger.log("=" * 50)
        
        return agg_df
        
    except FileNotFoundError as e:
        logger.log(f"\n文件未找到: {e}")
        logger.log("请检查 45G工参、路测网格、区域、网格、乡镇 等目录与文件。")
        raise
    except Exception as e:
        logger.log(f"\n处理数据时发生错误: {e}")
        logger.log(traceback.format_exc())
        raise
    finally:
        aggregator.close()


def run_physical_table_sector_fix(
    input_path: str,
    output_dir: str | None = None,
    auto_fix: bool = True,
    log_callback: LogCallback | None = None,
) -> dict:
    """检测并修正物理表扇区冲突
    
    Args:
        input_path: 物理表Excel文件路径
        output_dir: 输出目录，默认为输入文件所在目录
        auto_fix: 是否自动修正冲突
        log_callback: 日志回调函数
        
    Returns:
        包含 fixed_df, conflict_df, fix_df 的字典
    """
    logger = GuiLogger(log_callback)
    progress = GuiProgress(progress_callback, logger)
    
    if not PHYSICAL_TABLE_AVAILABLE:
        logger.log("错误: 物理表模块未安装，无法运行扇区修正功能")
        return {}
    
    progress.update(15, f"读取物理表: {os.path.basename(input_path)}")
    df = pd.read_excel(input_path)
    
    progress.update(40, "检测扇区冲突...")
    conflicts = detect_sector_conflicts(df)
    
    if output_dir is None:
        output_dir = os.path.dirname(input_path) or "."
    
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    
    result = {
        "conflicts": conflicts,
        "fixed_df": df.copy(),
        "conflict_df": pd.DataFrame(),
        "fix_df": pd.DataFrame(),
    }
    
    if conflicts.empty:
        progress.update(100, "未发现扇区冲突")
        logger.log("未发现扇区冲突")
        return result
    
    logger.log(f"发现 {len(conflicts)} 条冲突记录")
    
    if auto_fix:
        progress.update(60, "开始自动修正扇区冲突...")
        fixed_df, conflict_df, fix_df = auto_fix_sector_conflicts(df)
        progress.update(85, "保存修正结果...")
        
        # 保存结果
        out_path = os.path.join(output_dir, f"{base_name}-已修正.xlsx")
        fixed_df.to_excel(out_path, index=False)
        logger.log(f"已保存修正结果: {out_path}")
        
        if not conflict_df.empty:
            conflict_out = os.path.join(output_dir, f"{base_name}-扇区冲突明细.xlsx")
            conflict_df.to_excel(conflict_out, index=False)
            logger.log(f"已保存冲突明细: {conflict_out}")
        
        if not fix_df.empty:
            fix_out = os.path.join(output_dir, f"{base_name}-扇区修正明细.xlsx")
            fix_df.to_excel(fix_out, index=False)
            logger.log(f"已保存修正明细: {fix_out}")
        
        result["fixed_df"] = fixed_df
        result["conflict_df"] = conflict_df
        result["fix_df"] = fix_df
        progress.update(100, f"修正完成：冲突 {len(conflict_df)} 条，修正 {len(fix_df)} 条")
        
    return result


def get_data_file_status() -> dict:
    """扫描 data/ 目录，返回容量表与物理表源文件就绪状态。"""
    capacity: list[dict] = []
    capacity_ready = 0
    for key, pattern in FILE_PATTERNS.items():
        files = pick_matching_files(pattern)
        optional = key == "cog_coverage"
        found = len(files) > 0
        if found or optional:
            if found:
                capacity_ready += 1
        capacity.append(
            {
                "key": key,
                "pattern": pattern,
                "found": found,
                "optional": optional,
                "files": [f.name for f in files],
                "count": len(files),
            }
        )

    physical: list[dict] = []
    physical_ready = 0
    for key, pattern in PHYSICAL_FILE_PATTERNS.items():
        files = sorted(DATA_DIR.glob(pattern))
        files = [f for f in files if not f.name.startswith(".~")]
        found = len(files) > 0
        if found:
            physical_ready += 1
        physical.append(
            {
                "key": key,
                "pattern": pattern,
                "found": found,
                "optional": False,
                "files": [f.name for f in files],
                "count": len(files),
            }
        )

    required_capacity = len([k for k in FILE_PATTERNS if k != "cog_coverage"])
    required_found = sum(
        1 for item in capacity if item["found"] and not item["optional"]
    )

    return {
        "data_dir": str(DATA_DIR),
        "capacity": capacity,
        "physical": physical,
        "capacity_ready": capacity_ready,
        "capacity_total": len(FILE_PATTERNS),
        "capacity_required_ready": required_found == required_capacity,
        "physical_ready": physical_ready,
        "physical_total": len(PHYSICAL_FILE_PATTERNS),
        "physical_all_ready": physical_ready == len(PHYSICAL_FILE_PATTERNS),
    }


def list_output_files() -> list[dict]:
    """列出项目根目录下可下载的结果 Excel。"""
    patterns = [
        "合成_容量表_*.xlsx",
        "容量表_45G_*.xlsx",
        "物理表汇总结果.xlsx",
        "物理表汇总结果-*.xlsx",
        "低效小区结果.xlsx",
        "4G日监控_零低流量风险小区_*.xlsx",
        "*-扇区冲突明细.xlsx",
        "*-扇区修正明细.xlsx",
        "*-已修正.xlsx",
    ]
    seen: set[Path] = set()
    results: list[dict] = []
    for pattern in patterns:
        for path in sorted(BASE_DIR.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True):
            if path in seen or not path.is_file():
                continue
            seen.add(path)
            results.append(
                {
                    "name": path.name,
                    "size": path.stat().st_size,
                    "mtime": path.stat().st_mtime,
                }
            )
    return results


def run_low_efficiency_pipeline(
    progress_callback: ProgressCallback | None = None,
    log_callback: LogCallback | None = None,
) -> Path:
    """仅生成低效小区结果（含容量中间表计算，但不写出容量表）。"""
    logger = GuiLogger(log_callback)
    logger.log("=" * 50)
    logger.log("低效小区分析工具启动")
    logger.log("=" * 50)
    progress = GuiProgress(progress_callback, logger)
    init_db()
    conn = get_db_connection()
    try:
        progress.update(3, "开始导入数据到数据库")
        logger.log("准备扫描并导入源文件到 DuckDB 数据库")
        load_sources_to_db(conn, logger)
        table_5g, table_4g, _output_paths = _build_capacity_tables(conn, logger, progress)
        logger.log("开始合并 45G 总表")
        start = time.perf_counter()
        table_45g = build_45g_table(table_5g, table_4g)
        elapsed = time.perf_counter() - start
        progress.update(60, f"45G总表生成完成，共 {len(table_45g)} 条")
        logger.log(f"45G 总表生成完成 (耗时 {elapsed:.1f}s)")
        logger.log("开始生成低效小区结果")
        loweff_5g, loweff_4g, loweff_4g_full, loweff_summary = build_low_efficiency_table(
            table_5g, table_4g, table_45g
        )
        logger.log(
            f"低效小区结果生成完成，5G {len(loweff_5g)} 条，4G {len(loweff_4g)} 条，"
            f"全量4G评估 {len(loweff_4g_full)} 条"
        )
        with pd.ExcelWriter(LOWEFF_OUTPUT_PATH) as writer:
            loweff_5g.to_excel(writer, index=False, sheet_name="5G低效明细")
            loweff_4g.to_excel(writer, index=False, sheet_name="4G低效明细")
            loweff_4g_full.to_excel(writer, index=False, sheet_name="全量4G小区评估")
            loweff_summary.to_excel(writer, index=False, sheet_name="统计汇总")
        logger.log(f"写出 {LOWEFF_OUTPUT_PATH.name} 完成")
        progress.update(100, f"已生成: {LOWEFF_OUTPUT_PATH.name}")
        logger.log("=" * 50)
        return LOWEFF_OUTPUT_PATH
    finally:
        conn.close()
        try:
            if DB_PATH.exists():
                DB_PATH.unlink()
                logger.log("临时数据库已清理")
        except OSError:
            pass

