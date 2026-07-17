#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
4G日监控 - 零低流量风险小区分析脚本
用法：
    将天级Excel文件放入 INPUT_DIR 目录，运行脚本即可生成监控报表。
    支持任意天数的数据，天数越多连续天数计算越准确。
    自动关联"问题小区问题归类.xlsx"中的已知问题。

监控规则：
  1) 当日零流量：当天流量 = 0
  2) 当日低流量：0 < 当天流量 < 0.1GB (100MB)
  3) 风险等级按连续零/低流量天数划分
"""

import os
import sys
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# ==================== 用户配置区 ====================
INPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "天流量")
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PREFIX = "4G日监控_零低流量风险小区"
PROBLEM_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "问题小区问题归类.xlsx")
LOW_THRESHOLD_GB = 0.1          # 低流量阈值：100MB = 0.1GB
SHEET_NAME = None               # 设为 None 自动识别第一个sheet；也可固定为 "重要场景-天"
# ===================================================


def find_excel_files(directory):
    """查找目录下所有xlsx文件并按文件名排序"""
    if not os.path.isdir(directory):
        print(f"[错误] 目录不存在: {directory}")
        sys.exit(1)
    files = sorted([f for f in os.listdir(directory) if f.lower().endswith('.xlsx')])
    if not files:
        print(f"[错误] 目录下未找到xlsx文件: {directory}")
        sys.exit(1)
    return [os.path.join(directory, f) for f in files]


def load_and_merge(file_paths):
    """读取并合并所有天级文件"""
    all_dfs = []
    for fp in file_paths:
        xl = pd.ExcelFile(fp)
        sheet = SHEET_NAME if SHEET_NAME else xl.sheet_names[0]
        df = pd.read_excel(fp, sheet_name=sheet)
        df['__源文件__'] = os.path.basename(fp)
        all_dfs.append(df)
        print(f"  加载: {os.path.basename(fp)} => {len(df)} 行, {df['CGI'].nunique()} 个小区")
    merged = pd.concat(all_dfs, ignore_index=True)
    return merged


def load_problems(problem_file):
    """读取问题小区归类表"""
    if not os.path.exists(problem_file):
        print(f"[警告] 问题小区文件不存在: {problem_file}")
        return None
    xl = pd.ExcelFile(problem_file)
    sheet = xl.sheet_names[0]
    df = pd.read_excel(problem_file, sheet_name=sheet)
    # 处理问题列：将0或空值转为空字符串
    df['问题'] = df['问题'].apply(lambda x: '' if pd.isna(x) or str(x).strip() == '0' else str(x).strip())
    print(f"  加载问题小区: {problem_file} => {len(df)} 行, 有效问题数: {(df['问题'] != '').sum()}")
    return df[['CGI', '问题']]


def calc_consecutive(row, date_cols, mode='zero'):
    """从最新日期往前计算连续满足条件的天数"""
    count = 0
    for d in reversed(date_cols):
        v = row.get(d)
        if pd.isna(v):
            break
        if mode == 'zero':
            if v == 0:
                count += 1
            else:
                break
        elif mode == 'low':
            if v > 0 and v < LOW_THRESHOLD_GB:
                count += 1
            else:
                break
    return count


def risk_level(zd, ld):
    """根据连续零/低流量天数判定风险等级"""
    if zd >= 7:
        return '严重'
    if zd >= 5 or ld >= 7:
        return '高危'
    if zd >= 3 or ld >= 5:
        return '中危'
    if zd >= 1 or ld >= 3:
        return '预警'
    if ld >= 1:
        return '关注'
    return '正常'


def main():
    print("=" * 60)
    print("4G日监控 - 零低流量风险小区分析")
    print("=" * 60)

    # 1. 读取天级数据
    print(f"\n[1/5] 读取数据目录: {INPUT_DIR}")
    file_paths = find_excel_files(INPUT_DIR)
    raw_df = load_and_merge(file_paths)
    print(f"\n  合并完成: {len(raw_df)} 行, {raw_df['CGI'].nunique()} 个唯一小区")

    # 2. 读取问题小区
    print("\n[2/5] 读取问题小区归类...")
    problem_df = load_problems(PROBLEM_FILE)

    # 3. 数据预处理
    print("\n[3/5] 数据预处理...")
    raw_df['日期'] = pd.to_datetime(raw_df['记录开始时间']).dt.date
    raw_df['流量_GB'] = pd.to_numeric(raw_df['日4G流量（GB）'], errors='coerce')

    dates = sorted(raw_df['日期'].unique())
    monitor_date = dates[-1]
    # 生成带评估日期时间戳的输出文件名
    output_file = os.path.join(OUTPUT_DIR, f"{OUTPUT_PREFIX}_{monitor_date}.xlsx")
    print(f"  数据日期范围: {dates[0]} 至 {dates[-1]}, 共 {len(dates)} 天")
    print(f"  监控日期(最新): {monitor_date}")

    # 4. 透视：每行一个小区，每列一天流量
    print("\n[4/5] 构建透视表并计算风险...")
    pivot = raw_df.pivot_table(
        index=['CGI', '小区名称', '所属地市'],
        columns='日期',
        values='流量_GB',
        aggfunc='first'
    ).reset_index()
    pivot.columns.name = None

    date_cols = [str(d) for d in dates]
    for d in dates:
        pivot = pivot.rename(columns={d: str(d)})

    # 当日流量与状态
    pivot['当日流量_GB'] = pivot[str(monitor_date)]
    pivot['当日状态'] = pivot['当日流量_GB'].apply(
        lambda x: '零流量' if pd.isna(x) or x == 0 else ('低流量' if x < LOW_THRESHOLD_GB else '正常')
    )

    # 连续天数
    pivot['连续零流量天数'] = pivot.apply(lambda r: calc_consecutive(r, date_cols, 'zero'), axis=1)
    pivot['连续低流量天数'] = pivot.apply(lambda r: calc_consecutive(r, date_cols, 'low'), axis=1)
    pivot['连续零低流量天数'] = pivot[['连续零流量天数', '连续低流量天数']].max(axis=1)

    # 风险等级
    pivot['风险等级'] = pivot.apply(lambda r: risk_level(r['连续零流量天数'], r['连续低流量天数']), axis=1)

    # 关联问题小区
    if problem_df is not None:
        pivot = pivot.merge(problem_df, on='CGI', how='left')
        pivot['问题'] = pivot['问题'].fillna('')
    else:
        pivot['问题'] = ''

    total = len(pivot)
    zero_today = len(pivot[pivot['当日状态'] == '零流量'])
    low_today = len(pivot[pivot['当日状态'] == '低流量'])
    zero_low_today = zero_today + low_today
    ratio = zero_low_today / total if total else 0
    risk_counts = pivot['风险等级'].value_counts().to_dict()
    with_problem = (pivot['问题'] != '').sum()

    print(f"  小区总数: {total}")
    print(f"  当日零流量: {zero_today}, 当日低流量: {low_today}, 合计: {zero_low_today}")
    print(f"  当日零低流量占比: {ratio:.4%}")
    print(f"  风险分布: {dict(sorted(risk_counts.items(), key=lambda kv: -kv[1]))}")
    print(f"  已关联已知问题: {with_problem} 个小区")

    # 5. 输出Excel
    print("\n[5/5] 生成监控报表...")
    wb = Workbook()

    # --- Sheet1: 监控汇总 ---
    ws1 = wb.active
    ws1.title = "监控汇总"
    summary = [
        ["监控指标", "数值"],
        ["监控日期", str(monitor_date)],
        ["数据范围", f"{dates[0]} 至 {dates[-1]}（共{len(dates)}天）"],
        ["小区总数", total],
        ["当日零流量小区数", zero_today],
        ["当日低流量小区数", low_today],
        ["当日零低流量小区数", zero_low_today],
        ["当日零低流量占比", f"{ratio:.4%}"],
        ["已关联已知问题小区数", with_problem],
        ["", ""],
        ["风险等级", "小区数"],
    ]
    for lv in ['严重', '高危', '中危', '预警', '关注', '正常']:
        summary.append([lv, risk_counts.get(lv, 0)])

    for r_idx, row in enumerate(summary, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws1.cell(r_idx, c_idx, val)
            if r_idx == 1 or (r_idx == 11 and c_idx == 1):
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="366092")
                cell.alignment = Alignment(horizontal="center")
            elif c_idx == 1:
                cell.font = Font(bold=True)
            elif r_idx == 8 and c_idx == 2:
                cell.font = Font(bold=True, color="C00000")
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 22

    # --- Sheet2: 风险小区明细 ---
    ws2 = wb.create_sheet("风险小区明细")
    risk_df = pivot[pivot['当日状态'].isin(['零流量', '低流量'])].copy()
    risk_df = risk_df.sort_values(['风险等级', '连续零低流量天数'], ascending=[True, False])

    out_cols = ['CGI', '小区名称', '所属地市', '当日流量_GB', '当日状态',
                '连续零流量天数', '连续低流量天数', '连续零低流量天数', '风险等级', '问题']
    risk_out = risk_df[out_cols].copy()
    risk_out.columns = ['CGI', '小区名称', '所属地市', '当日流量(GB)', '当日状态',
                        '连续零流量天数', '连续低流量天数', '连续零低流量天数', '风险等级', '已知问题']

    risk_colors = {
        '严重': ('C00000', 'FFFFFF'),
        '高危': ('FF0000', 'FFFFFF'),
        '中危': ('E36C0A', 'FFFFFF'),
        '预警': ('FFC000', '000000'),
        '关注': ('FFFF00', '000000'),
        '正常': ('FFFFFF', '000000'),
    }

    for r_idx, row in enumerate(dataframe_to_rows(risk_out, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws2.cell(r_idx, c_idx, val)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="366092")
                cell.alignment = Alignment(horizontal="center")
            else:
                if c_idx == 9:  # 风险等级列
                    bg, fg = risk_colors.get(val, ('FFFFFF', '000000'))
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(bold=True, color=fg)
                elif c_idx == 10 and val:  # 已知问题列且有值
                    cell.font = Font(color="C00000")
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
                elif r_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F7F9FC")

    for col in ['A', 'B', 'C']:
        ws2.column_dimensions[col].width = 26
    for col in ['D', 'E', 'F', 'G', 'H', 'I']:
        ws2.column_dimensions[col].width = 16
    ws2.column_dimensions['J'].width = 40

    # --- Sheet3: 全量监控明细 ---
    ws3 = wb.create_sheet("全量监控明细")
    all_out = pivot[['CGI', '小区名称', '所属地市'] + date_cols +
                    ['当日流量_GB', '当日状态', '连续零流量天数', '连续低流量天数',
                     '连续零低流量天数', '风险等级', '问题']].copy()
    all_out.columns = (['CGI', '小区名称', '所属地市'] +
                       [f'{d}流量(GB)' for d in dates] +
                       ['当日流量(GB)', '当日状态', '连续零流量天数', '连续低流量天数',
                        '连续零低流量天数', '风险等级', '已知问题'])

    for r_idx, row in enumerate(dataframe_to_rows(all_out, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws3.cell(r_idx, c_idx, val)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="366092")
                cell.alignment = Alignment(horizontal="center")
            else:
                if c_idx == len(row) - 1:  # 风险等级列
                    bg, fg = risk_colors.get(val, ('FFFFFF', '000000'))
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(bold=True, color=fg)
                elif c_idx == len(row) and val:  # 已知问题列且有值
                    cell.font = Font(color="C00000")
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
                elif r_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F7F9FC")

    for col in ['A', 'B', 'C']:
        ws3.column_dimensions[col].width = 26
    for c in range(4, 4 + len(dates)):
        col_letter = chr(64 + c) if c <= 26 else 'A' + chr(64 + c - 26)
        ws3.column_dimensions[col_letter].width = 16
    for col in ['K', 'L', 'M', 'N', 'O']:
        ws3.column_dimensions[col].width = 16
    ws3.column_dimensions[get_column_letter(len(all_out.columns))].width = 40

    wb.save(output_file)
    print(f"\n  报表已保存: {output_file}")
    print("\n" + "=" * 60)
    print("监控完成!")
    print("=" * 60)


if __name__ == '__main__':
    main()
